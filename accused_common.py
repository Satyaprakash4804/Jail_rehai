"""
accused_common.py
=================
अभियुक्त प्रबंधन का मुख्य तर्क (core logic)

Handles:
  • Excel/CSV upload → FIR case + accused parsing
  • Accused deduplication (fuzzy name + father match)
  • Accused list with FIR filter
  • Accused detail with all linked FIRs and acts
  • Manual FIR case creation
"""

import io
import re
import base64
import logging
from difflib import SequenceMatcher
from datetime import datetime

from flask import render_template, request, redirect, url_for, session, flash, send_file
from db import get_connection
from utils import log_activity, upload_image, paginate_query, upload_document, send_bail_notification

logger = logging.getLogger(__name__)

# ── Helpers ───────────────────────────────────────────────────────────────────

def _district(): return session.get('district')
def _uid():      return session.get('user_id')
def _role():     return session.get('role')
def _bp():       return 'admin' if _role() == 'admin' else 'super'


# Values that mean "no real name given" — an Accused record must NEVER be
# created for these. They only ever appear as a father's-name placeholder,
# never as a standalone accused entry.
UNKNOWN_MARKERS = {
    'अज्ञात', 'अज्ञात व्यक्ति', 'unknown', 'na', 'n/a', 'n.a.', 'nil',
    '-', '--', '.', 'none', 'not known', 'unnamed',
}


def is_unknown_name(name: str) -> bool:
    """True if `name` is empty or is one of the 'unknown/अज्ञात' markers."""
    if not name:
        return True
    cleaned = re.sub(r'[.\s]+', ' ', name).strip().lower()
    return (not cleaned) or (cleaned in UNKNOWN_MARKERS)


def normalize_name(name: str) -> str:
    """
    Normalize a Hindi/English name for dedup comparison.
    Removes extra spaces, converts to lowercase,
    strips common honorifics (श्री, श्रीमती, etc.)
    """
    if not name:
        return ''
    n = name.strip()
    # Remove common prefixes
    for prefix in ['श्री ', 'श्रीमती ', 'Mr. ', 'Mrs. ', 'Smt. ', 'Shri ']:
        if n.startswith(prefix):
            n = n[len(prefix):]
    return re.sub(r'\s+', ' ', n).strip().lower()


def fuzzy_match(a: str, b: str, threshold: float = 0.82) -> bool:
    """
    Returns True if two normalized strings are similar enough
    to be considered the same person.
    Uses SequenceMatcher (works for Devanagari Unicode strings).
    """
    if not a or not b:
        return False
    ratio = SequenceMatcher(None, a, b).ratio()
    return ratio >= threshold


def parse_accused_list(raw: str) -> list:
    """
    Parse comma-separated accused string like:
    'अंकित पाल s/o रामअवतार, अभिषेक विश्वकर्मा s/o शिवकुमार'
    into list of dicts: [{'name': ..., 'fathers_name': ...}, ...]

    Handles 's/o', 'S/O', 'पुत्र', 'w/o', 'd/o' separators.

    IMPORTANT: entries with no real name — blank, or a plain "अज्ञात"/
    "unknown" marker — are DROPPED here and never become an Accused
    record. Only the father's-name side may legitimately be 'अज्ञात'
    when the name itself is real.
    """
    if not raw:
        return []
    results = []
    # Split on comma but not commas inside parentheses
    parts = [p.strip() for p in raw.split(',') if p.strip()]
    for part in parts:
        # Match patterns: NAME s/o FATHER or NAME पुत्र FATHER
        m = re.split(r'\s+(?:s/o|S/O|S\/O|पुत्र|w/o|W/O|d/o|D/O|पुत्री)\s+', part, maxsplit=1)
        if len(m) == 2:
            name         = m[0].strip()
            fathers_name = m[1].strip()
        else:
            # No s/o found — take full string as name, father unknown
            name         = part.strip()
            fathers_name = 'अज्ञात'

        # Never create an accused dataset entry for an unnamed/अज्ञात person
        if is_unknown_name(name):
            continue
        if not fathers_name or is_unknown_name(fathers_name):
            fathers_name = 'अज्ञात'

        results.append({'name': name, 'fathers_name': fathers_name})
    return results


def find_or_create_accused(cursor, name: str, fathers_name: str, created_by: int):
    """
    Find an existing accused record using fuzzy match on
    normalized name + father name.
    If not found, create a new record.

    Returns (accused_id, was_created) — was_created is False when an
    existing (deduplicated) record was matched, so callers can report
    accurate "created vs linked" counts instead of double-counting.

    Never creates a record for an unknown/अज्ञात/blank name — returns
    (None, False) in that case so the caller can skip it entirely.
    """
    if is_unknown_name(name):
        return None, False

    name_norm    = normalize_name(name)
    father_norm  = normalize_name(fathers_name)

    # Exact normalized match first (fast)
    cursor.execute("""
        SELECT id, name, fathers_name, name_normalized, fathers_normalized
        FROM accused
        WHERE name_normalized = %s AND fathers_normalized = %s
        LIMIT 1
    """, (name_norm, father_norm))
    row = cursor.fetchone()
    if row:
        return row['id'], False

    # Fuzzy match — check candidates with same first character
    first_char = name_norm[0] if name_norm else ''
    cursor.execute("""
        SELECT id, name_normalized, fathers_normalized
        FROM accused
        WHERE name_normalized LIKE %s
    """, (first_char + '%',))
    candidates = cursor.fetchall()

    for cand in candidates:
        if (fuzzy_match(name_norm, cand['name_normalized']) and
                fuzzy_match(father_norm, cand['fathers_normalized'])):
            logger.info(
                f"[Dedup] Matched '{name}' s/o '{fathers_name}' "
                f"→ accused_id={cand['id']}"
            )
            return cand['id'], False

    # Not found — create new accused record
    cursor.execute("""
        INSERT INTO accused
            (name, name_normalized, fathers_name, fathers_normalized, created_by)
        VALUES (%s, %s, %s, %s, %s)
    """, (name, name_norm, fathers_name, father_norm, created_by))
    new_id = cursor.lastrowid
    logger.info(f"[Accused] Created new accused_id={new_id}: {name} s/o {fathers_name}")
    return new_id, True


def upsert_accused_fir(cursor, accused_id: int, fir_id: int,
                       in_total: bool, in_fir: bool,
                       in_arrested: bool, in_cs: bool):
    """
    Insert or update the accused_fir junction row.
    Uses OR logic so flags only grow (never reset).
    """
    cursor.execute("""
        INSERT INTO accused_fir
            (accused_id, fir_id, in_total_accused, in_fir_accused, in_arrested, in_cs_accused)
        VALUES (%s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE
            in_total_accused = in_total_accused OR VALUES(in_total_accused),
            in_fir_accused   = in_fir_accused   OR VALUES(in_fir_accused),
            in_arrested      = in_arrested       OR VALUES(in_arrested),
            in_cs_accused    = in_cs_accused     OR VALUES(in_cs_accused)
    """, (accused_id, fir_id,
          1 if in_total else 0,
          1 if in_fir else 0,
          1 if in_arrested else 0,
          1 if in_cs else 0))


# ── Excel Upload ──────────────────────────────────────────────────────────────

def upload_accused_excel(role='admin'):
    """
    Handle Excel upload for FIR case + accused creation.
    Reads the UP Police format Excel with columns:
    क्र., जनपद, थाना, FIR संख्या, धारा,
    कुल अभियुक्त, FIR में अभियुक्त, गिरफ्तार, आरोप पत्र, वादी, स्थिति
    """
    tmpl = f'{_bp()}/upload_accused.html'

    if request.method == 'POST':
        file = request.files.get('excel_file')
        if not file or not file.filename:
            flash('कृपया Excel फ़ाइल चुनें।', 'danger')
            return render_template(tmpl)

        fname = file.filename.lower()
        if not (fname.endswith('.xlsx') or fname.endswith('.xls')):
            flash('केवल .xlsx या .xls फ़ाइल अपलोड करें।', 'danger')
            return render_template(tmpl)

        try:
            import openpyxl
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
        except Exception as e:
            flash(f'Excel फ़ाइल पढ़ने में त्रुटि: {e}', 'danger')
            return render_template(tmpl)

        conn = get_connection()
        cursor = conn.cursor(dictionary=True)
        district_session = _district()

        fir_success = 0
        accused_created = 0   # brand-new accused records
        accused_linked  = 0   # accused-FIR links created/updated (existing or new accused)
        accused_skipped = 0   # unknown/अज्ञात/blank entries dropped
        failed = []

        # Skip header row (row 1)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue

            try:
                # Unpack columns
                (sr_no, district, thana, fir_number, acts,
                 total_raw, fir_raw, arrested_raw, cs_raw,
                 complainant, status) = (list(row) + [None] * 11)[:11]

                # Use session district if district column is empty
                district = str(district or '').strip() or district_session
                thana       = str(thana       or '').strip()
                fir_number  = str(fir_number  or '').strip()
                acts        = str(acts        or '').strip()
                total_raw   = str(total_raw   or '').strip()
                fir_raw     = str(fir_raw     or '').strip()
                arrested_raw= str(arrested_raw or '').strip()
                cs_raw      = str(cs_raw      or '').strip()
                complainant = str(complainant or '').strip()
                status      = str(status      or '').strip()

                if not fir_number or not thana:
                    failed.append(f"पंक्ति {row_idx}: FIR संख्या या थाना खाली")
                    continue

                # ── Create or get FIR case ────────────────────────────────
                cursor.execute("""
                    INSERT INTO fir_cases
                        (district, thana, fir_number, acts,
                         total_accused_raw, fir_accused_raw,
                         arrested_accused_raw, cs_accused_raw,
                         complainant, status, created_by)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    ON DUPLICATE KEY UPDATE
                        acts=VALUES(acts),
                        total_accused_raw=VALUES(total_accused_raw),
                        fir_accused_raw=VALUES(fir_accused_raw),
                        arrested_accused_raw=VALUES(arrested_accused_raw),
                        cs_accused_raw=VALUES(cs_accused_raw),
                        complainant=VALUES(complainant),
                        status=VALUES(status)
                """, (district, thana, fir_number, acts,
                      total_raw, fir_raw, arrested_raw, cs_raw,
                      complainant, status, _uid()))

                # Get fir_id
                cursor.execute("""
                    SELECT id FROM fir_cases
                    WHERE district=%s AND thana=%s AND fir_number=%s
                """, (district, thana, fir_number))
                fir_row = cursor.fetchone()
                if not fir_row:
                    continue
                fir_id = fir_row['id']
                fir_success += 1

                # ── Parse and process each accused column ─────────────────
                # parse_accused_list() already drops blank/अज्ञात/unknown
                # names, so every set below only ever contains real people.
                total_set    = set(normalize_name(a['name']) for a in parse_accused_list(total_raw))
                fir_set      = set(normalize_name(a['name']) for a in parse_accused_list(fir_raw))
                arrested_set = set(normalize_name(a['name']) for a in parse_accused_list(arrested_raw))
                cs_set       = set(normalize_name(a['name']) for a in parse_accused_list(cs_raw))

                # Process all unique accused from total column
                all_accused = parse_accused_list(total_raw)
                # Also add any in fir_raw not already in total (dedup by normalized name)
                seen_norms = set(total_set)
                for a in parse_accused_list(fir_raw):
                    n = normalize_name(a['name'])
                    if n not in seen_norms:
                        all_accused.append(a)
                        seen_norms.add(n)

                linked_this_fir = 0
                for acc in all_accused:
                    acc_name   = acc['name']
                    acc_father = acc['fathers_name']
                    acc_norm   = normalize_name(acc_name)

                    accused_id, was_created = find_or_create_accused(
                        cursor, acc_name, acc_father, _uid()
                    )
                    if accused_id is None:
                        accused_skipped += 1
                        continue
                    if was_created:
                        accused_created += 1

                    in_total    = acc_norm in total_set
                    in_fir      = acc_norm in fir_set
                    in_arrested = acc_norm in arrested_set
                    in_cs       = acc_norm in cs_set

                    upsert_accused_fir(
                        cursor, accused_id, fir_id,
                        in_total, in_fir, in_arrested, in_cs
                    )
                    accused_linked += 1
                    linked_this_fir += 1

                # Update total count — only counts real, linked accused
                cursor.execute(
                    "UPDATE fir_cases SET total_accused_count=%s WHERE id=%s",
                    (linked_this_fir, fir_id)
                )

            except Exception as e:
                logger.error(f"Row {row_idx} error: {e}")
                failed.append(f"पंक्ति {row_idx}: {str(e)[:100]}")
                continue

        conn.commit()
        cursor.close()
        conn.close()

        log_activity(_uid(), _role(),
                     f"Excel upload: {fir_success} FIR, {accused_created} नए अभियुक्त, "
                     f"{accused_linked} लिंक, {accused_skipped} अज्ञात/खाली छोड़े गए",
                     ip=request.remote_addr)

        flash(
            f"✅ {fir_success} FIR मामले | {accused_created} नए अभियुक्त बनाए गए | "
            f"{accused_linked} अभियुक्त-FIR लिंक किए गए"
            + (f" | {accused_skipped} अज्ञात/खाली नाम छोड़े गए" if accused_skipped else "")
            + f" | असफल पंक्तियाँ: {len(failed)}",
            'success' if not failed else 'warning'
        )
        for msg in failed[:10]:
            flash(msg, 'warning')

        return redirect(url_for(f'{_bp()}.fir_list'))

    return render_template(tmpl)


# ── FIR List ──────────────────────────────────────────────────────────────────

def get_fir_list(role='admin'):
    """List all FIR cases with filters: thana, fir_number, search"""
    district  = _district()
    page      = int(request.args.get('page', 1))
    per_page  = int(request.args.get('per_page', 25))
    search    = request.args.get('search', '').strip()
    thana_f   = request.args.get('thana', '').strip()
    fir_f     = request.args.get('fir', '').strip()

    conditions = ["f.district=%s"]
    params     = [district]

    if thana_f:
        conditions.append("f.thana LIKE %s")
        params.append(f'%{thana_f}%')
    if fir_f:
        conditions.append("f.fir_number LIKE %s")
        params.append(f'%{fir_f}%')
    if search:
        conditions.append("(f.fir_number LIKE %s OR f.thana LIKE %s OR f.complainant LIKE %s OR f.acts LIKE %s)")
        like = f'%{search}%'
        params += [like, like, like, like]

    where  = " AND ".join(conditions)
    conn   = get_connection()
    cursor = conn.cursor(dictionary=True)

    base_q = f"""
        SELECT f.*,
               (SELECT COUNT(*) FROM accused_fir af WHERE af.fir_id=f.id) AS accused_count
        FROM fir_cases f
        WHERE {where}
        ORDER BY f.created_at DESC
    """
    rows, total, total_pages = paginate_query(cursor, base_q, params, page, per_page)

    # Get unique thana list for filter dropdown
    cursor.execute(
        "SELECT DISTINCT thana FROM fir_cases WHERE district=%s ORDER BY thana",
        (district,)
    )
    thanas = [r['thana'] for r in cursor.fetchall()]
    cursor.close()
    conn.close()

    tmpl = f'{role}/fir_list.html' if role == 'admin' else 'super/fir_list.html'
    return render_template(
        tmpl,
        firs=rows, page=page, total=total, total_pages=total_pages,
        per_page=per_page, search=search, thana_filter=thana_f,
        fir_filter=fir_f, thanas=thanas
    )


# ── FIR Detail ────────────────────────────────────────────────────────────────

def get_fir_detail(fir_id: int):
    district = _district()
    conn     = get_connection()
    cursor   = conn.cursor(dictionary=True)

    cursor.execute(
        "SELECT * FROM fir_cases WHERE id=%s AND district=%s",
        (fir_id, district)
    )
    fir = cursor.fetchone()
    if not fir:
        flash('FIR मामला नहीं मिला।', 'danger')
        cursor.close(); conn.close()
        return redirect(url_for(f'{_bp()}.fir_list'))

    # All accused in this FIR with their flags
    cursor.execute("""
        SELECT a.id, a.name, a.fathers_name, a.photo_url, a.profile_status,
               af.in_total_accused, af.in_fir_accused, af.in_arrested, af.in_cs_accused,
               (SELECT COUNT(*) FROM accused_fir af2 WHERE af2.accused_id=a.id) AS fir_count
        FROM accused_fir af
        JOIN accused a ON a.id = af.accused_id
        WHERE af.fir_id = %s
        ORDER BY a.name
    """, (fir_id,))
    accused_list = cursor.fetchall()

    cursor.close(); conn.close()
    tmpl = f'{_bp()}/fir_detail.html'
    return render_template(tmpl, fir=fir, accused_list=accused_list)


# ── Accused List ──────────────────────────────────────────────────────────────

def get_accused_list(role='admin'):
    """
    List all accused with filters.
    Can filter by thana or FIR number (cross-reference view).
    """
    district  = _district()
    page      = int(request.args.get('page', 1))
    per_page  = int(request.args.get('per_page', 25))
    search    = request.args.get('search', '').strip()
    thana_f   = request.args.get('thana', '').strip()
    fir_f     = request.args.get('fir', '').strip()
    status_f  = request.args.get('status', '').strip()

    conn   = get_connection()
    cursor = conn.cursor(dictionary=True)

    # Base: accused who appear in at least one FIR in this district
    conditions = ["f.district=%s"]
    params     = [district]

    if thana_f:
        conditions.append("f.thana LIKE %s")
        params.append(f'%{thana_f}%')
    if fir_f:
        conditions.append("f.fir_number LIKE %s")
        params.append(f'%{fir_f}%')
    if status_f in ('pending', 'complete'):
        conditions.append("a.profile_status=%s")
        params.append(status_f)
    if search:
        conditions.append("(a.name LIKE %s OR a.fathers_name LIKE %s)")
        like = f'%{search}%'
        params += [like, like]

    where = " AND ".join(conditions)

    base_q = f"""
        SELECT DISTINCT
            a.id, a.name, a.fathers_name, a.photo_url, a.profile_status,
            (SELECT COUNT(*) FROM accused_fir af2 WHERE af2.accused_id=a.id) AS fir_count,
            (SELECT GROUP_CONCAT(DISTINCT f2.thana ORDER BY f2.thana SEPARATOR ', ')
             FROM accused_fir af3 JOIN fir_cases f2 ON f2.id=af3.fir_id
             WHERE af3.accused_id=a.id AND f2.district=%s) AS thanas,
            MAX(af.in_arrested) AS ever_arrested
        FROM accused a
        JOIN accused_fir af ON af.accused_id = a.id
        JOIN fir_cases f ON f.id = af.fir_id
        WHERE {where}
        GROUP BY a.id
        ORDER BY a.name
    """
    # Add district param for subquery
    params_full = [district] + params

    rows, total, total_pages = paginate_query(cursor, base_q, params_full, page, per_page)

    # Thana dropdown
    cursor.execute(
        "SELECT DISTINCT thana FROM fir_cases WHERE district=%s ORDER BY thana",
        (district,)
    )
    thanas = [r['thana'] for r in cursor.fetchall()]
    cursor.close(); conn.close()

    tmpl = f'{role}/accused_list.html'
    return render_template(
        tmpl,
        accused=rows, page=page, total=total, total_pages=total_pages,
        per_page=per_page, search=search, thana_filter=thana_f,
        fir_filter=fir_f, status_filter=status_f, thanas=thanas
    )


# ── Accused Detail ────────────────────────────────────────────────────────────

def get_accused_detail(accused_id: int):
    district = _district()
    conn     = get_connection()
    cursor   = conn.cursor(dictionary=True)

    cursor.execute("SELECT * FROM accused WHERE id=%s", (accused_id,))
    accused = cursor.fetchone()
    if not accused:
        flash('अभियुक्त नहीं मिला।', 'danger')
        cursor.close(); conn.close()
        return redirect(url_for(f'{_bp()}.accused_list'))

    # All FIRs this accused is linked to (in the session district)
    cursor.execute("""
        SELECT f.*, af.in_total_accused, af.in_fir_accused,
               af.in_arrested, af.in_cs_accused
        FROM accused_fir af
        JOIN fir_cases f ON f.id = af.fir_id
        WHERE af.accused_id = %s AND f.district = %s
        ORDER BY f.fir_number
    """, (accused_id, district))
    firs = cursor.fetchall()

    # Photos
    cursor.execute("""
        SELECT * FROM accused_photos WHERE accused_id=%s ORDER BY uploaded_at DESC
    """, (accused_id,))
    photos = cursor.fetchall()

    # Bail eligibility: only accused arrested (in_arrested=1) in at least one FIR
    is_arrested   = any(f.get('in_arrested') for f in firs)
    arrest_firs   = [f for f in firs if f.get('in_arrested')]
    has_active_bail = bool(accused.get('bail_status') and accused['bail_status'] != 'none')

    # Full bail history (never filtered/deleted — always shown so repeated
    # arrests/FIRs for the same accused keep their complete jamanat record).
    # A bail record may now cover multiple FIRs (accused_bail_fir junction);
    # fir_number/thana stay as the *primary* FIR for backward-compatible
    # display, while all_firs/fir_count expose the full multi-FIR list.
    cursor.execute("""
        SELECT abh.*, abh.status AS bail_history_status, f.fir_number, f.thana, f.district,
               GROUP_CONCAT(DISTINCT CONCAT(f2.fir_number,'/',f2.thana)
                            ORDER BY f2.fir_number SEPARATOR ', ') AS all_firs,
               COUNT(DISTINCT f2.id) AS fir_count,
               u1.name AS approved_by_name, u2.name AS revoked_by_name
        FROM accused_bail_history abh
        JOIN fir_cases f ON f.id = abh.fir_id
        LEFT JOIN accused_bail_fir abf ON abf.bail_id = abh.id
        LEFT JOIN fir_cases f2 ON f2.id = abf.fir_id
        LEFT JOIN users u1 ON u1.id = abh.approved_by
        LEFT JOIN users u2 ON u2.id = abh.revoked_by
        WHERE abh.accused_id = %s
        GROUP BY abh.id
        ORDER BY abh.approved_at DESC
    """, (accused_id,))
    bail_history = cursor.fetchall()
    for b in bail_history:
        # Legacy rows created before accused_bail_fir existed have no
        # junction rows — fall back to the single primary FIR for display.
        if not b.get('all_firs'):
            b['all_firs'] = f"{b['fir_number']}/{b['thana']}"
            b['fir_count'] = 1
    cursor.close(); conn.close()

    tmpl = f'{_bp()}/accused_detail.html'
    return render_template(
        tmpl, accused=accused, firs=firs, photos=photos,
        is_arrested=is_arrested, arrest_firs=arrest_firs,
        has_active_bail=has_active_bail, bail_history=bail_history
    )


# ── Accused Bail (जमानत) — ONLY for arrested accused ──────────────────────────
# Terminology: the action is "Approve Bail" (जमानत स्वीकृत करें), not "Grant".
# History is permanent — every approval/revoke stays in accused_bail_history
# and is always shown on the accused's detail page, even across repeat FIRs.

def approve_accused_bail(accused_id: int, role='admin'):
    """
    GET  -> show approve-bail form (only if accused is arrested in >=1 FIR
            and has no currently active bail).
    POST -> create accused_bail_history row + update accused.bail_* (current).
    """
    bp = 'admin' if role == 'admin' else 'super'
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT * FROM accused WHERE id=%s", (accused_id,))
    accused = cursor.fetchone()
    if not accused:
        flash('अभियुक्त नहीं मिला।', 'danger')
        cursor.close(); conn.close()
        return redirect(url_for(f'{bp}.accused_list'))

    # Must be arrested in at least one FIR to be bail-eligible
    cursor.execute("""
        SELECT f.id, f.fir_number, f.thana, f.district
        FROM accused_fir af JOIN fir_cases f ON f.id = af.fir_id
        WHERE af.accused_id=%s AND af.in_arrested=1
        ORDER BY f.fir_number
    """, (accused_id,))
    arrest_firs = cursor.fetchall()

    if not arrest_firs:
        flash('जमानत केवल उन्हीं अभियुक्तों के लिए स्वीकृत की जा सकती है जो किसी FIR में गिरफ़्तार दर्शाए गए हैं।', 'danger')
        cursor.close(); conn.close()
        return redirect(url_for(f'{bp}.accused_detail', accused_id=accused_id))

    if accused.get('bail_status') and accused['bail_status'] != 'none':
        flash('इस अभियुक्त की जमानत पहले से सक्रिय है। नई जमानत स्वीकृत करने से पहले मौजूदा जमानत रद्द करें।', 'warning')
        cursor.close(); conn.close()
        return redirect(url_for(f'{bp}.accused_detail', accused_id=accused_id))

    if request.method == 'POST':
        # एक जमानत रिकॉर्ड में एक से अधिक गिरफ़्तारी FIR चुनने का समर्थन —
        # checkbox list "fir_ids"। पुराने single-select फ़ॉर्म से भी संगत
        # रहने हेतु "fir_id" fallback भी स्वीकार किया जाता है।
        fir_ids = request.form.getlist('fir_ids', type=int)
        if not fir_ids:
            legacy_fir_id = request.form.get('fir_id', type=int)
            if legacy_fir_id:
                fir_ids = [legacy_fir_id]
        bail_type   = request.form.get('bail_type', 'temporary')
        bail_start  = request.form.get('bail_start_date', '').strip() or None
        bail_end    = request.form.get('bail_end_date', '').strip() or None
        bail_remark = request.form.get('bail_remark', '').strip()
        bail_rating = int(request.form.get('bail_rating', 0) or 0)
        if bail_type == 'permanent':
            bail_end = None

        valid_fir_ids = {f['id'] for f in arrest_firs}
        fir_ids = [fid for fid in dict.fromkeys(fir_ids) if fid in valid_fir_ids]  # dedupe, keep order
        if not fir_ids:
            flash('कम से कम एक मान्य गिरफ़्तारी FIR चुनें।', 'danger')
            cursor.close(); conn.close()
            return render_template(f'{bp}/approve_accused_bail.html', accused=accused, arrest_firs=arrest_firs)
        fir_id = fir_ids[0]  # primary FIR — kept for backward-compat single-FIR column

        # जमानत की तिथियाँ भूतकाल में नहीं हो सकतीं (server-side check —
        # भले ही ब्राउज़र की date-picker पहले से ही पिछली तिथियाँ रोकती है)
        today = datetime.now().date()
        try:
            start_date_obj = datetime.strptime(bail_start, '%Y-%m-%d').date() if bail_start else None
        except ValueError:
            start_date_obj = None
        if not start_date_obj or start_date_obj < today:
            flash('जमानत शुरुआत तिथि आज या भविष्य की होनी चाहिए — पिछली तिथि मान्य नहीं है।', 'danger')
            cursor.close(); conn.close()
            return render_template(f'{bp}/approve_accused_bail.html', accused=accused, arrest_firs=arrest_firs)
        if bail_end:
            try:
                end_date_obj = datetime.strptime(bail_end, '%Y-%m-%d').date()
            except ValueError:
                end_date_obj = None
            if not end_date_obj or end_date_obj < start_date_obj:
                flash('जमानत समाप्ति तिथि शुरुआत तिथि से पहले नहीं हो सकती।', 'danger')
                cursor.close(); conn.close()
                return render_template(f'{bp}/approve_accused_bail.html', accused=accused, arrest_firs=arrest_firs)

        # ── अभियुक्त की फ़ोटो — या तो कैमरे से live geo-tagged कैप्चर, या
        # सीधे फ़ाइल अपलोड। दोनों में से कोई एक अनिवार्य है।
        # bail_photo_source बताता है कि फ़्रंटएंड ने कौन सा तरीका भेजा:
        #   'camera' → bail_photo_data (data:image base64, geo/time overlay
        #              पहले से छवि पर बेक किया हुआ)
        #   'upload' → bail_photo_upload (साधारण फ़ाइल, कोई GPS/समय overlay नहीं)
        photo_data     = request.form.get('bail_photo_data', '').strip()
        photo_source   = request.form.get('bail_photo_source', 'camera').strip()
        photo_lat_raw  = request.form.get('bail_photo_lat', '').strip()
        photo_lng_raw  = request.form.get('bail_photo_lng', '').strip()
        photo_time_raw = request.form.get('bail_photo_captured_at', '').strip()
        uploaded_photo_file = request.files.get('bail_photo_upload')

        bail_photo_url = bail_photo_public_id = None
        is_uploaded_photo = False

        if photo_data.startswith('data:image'):
            # ── कैमरा कैप्चर पथ ──────────────────────────────────────────
            bail_photo_url, bail_photo_public_id = upload_image(photo_data, folder='accused_bail_photos')
        elif uploaded_photo_file and uploaded_photo_file.filename:
            # ── मैनुअल फ़ाइल अपलोड पथ ────────────────────────────────────
            # कोई live GPS/समय overlay संभव नहीं — फ़ाइल जैसी है वैसी ही अपलोड होगी।
            is_uploaded_photo = True
            try:
                file_bytes = uploaded_photo_file.read()
                mime_type  = uploaded_photo_file.mimetype or 'image/jpeg'
                b64_data   = f"data:{mime_type};base64,{base64.b64encode(file_bytes).decode('utf-8')}"
                bail_photo_url, bail_photo_public_id = upload_image(b64_data, folder='accused_bail_photos')
            except Exception as e:
                logger.error(f"Bail photo upload (file) error: {e}")
                bail_photo_url = None
        else:
            flash('जमानत स्वीकृत करने से पहले अभियुक्त की फ़ोटो लेना या अपलोड करना अनिवार्य है।', 'danger')
            cursor.close(); conn.close()
            return render_template(f'{bp}/approve_accused_bail.html', accused=accused, arrest_firs=arrest_firs)

        if not bail_photo_url:
            flash('फ़ोटो अपलोड असफल रहा। कृपया पुनः प्रयास करें।', 'danger')
            cursor.close(); conn.close()
            return render_template(f'{bp}/approve_accused_bail.html', accused=accused, arrest_firs=arrest_firs)

        # अपलोड की गई फ़ोटो के लिए GPS/समय उपलब्ध नहीं — केवल कैमरा-कैप्चर के
        # लिए ब्राउज़र से मिले lat/lng/समय का उपयोग करें।
        if is_uploaded_photo:
            photo_lat_val = None
            photo_lng_val = None
            photo_time_val = datetime.now()
        else:
            try:
                photo_lat_val = float(photo_lat_raw) if photo_lat_raw else None
            except ValueError:
                photo_lat_val = None
            try:
                photo_lng_val = float(photo_lng_raw) if photo_lng_raw else None
            except ValueError:
                photo_lng_val = None
            try:
                photo_time_val = datetime.strptime(photo_time_raw, '%Y-%m-%dT%H:%M:%S') if photo_time_raw else datetime.now()
            except ValueError:
                photo_time_val = datetime.now()

        doc = request.files.get('bail_document')
        doc_url, doc_pub_id, doc_res_type = None, None, 'raw'
        if doc and doc.filename:
            doc_url, doc_pub_id, doc_res_type = upload_document(doc, folder='accused_bail_docs')

        # ── प्रोफ़ाइल फ़ोटो न हो तो जमानत के समय ली/अपलोड की गई फ़ोटो को ही
        # प्रोफ़ाइल फ़ोटो बना दें (कोई नया अपलोड नहीं — वही Cloudinary URL पुनः
        # उपयोग होता है), ताकि अभियुक्त कभी बिना फ़ोटो के न रहे।
        set_as_profile_photo = not accused.get('photo_url')
        if set_as_profile_photo:
            cursor.execute("UPDATE accused_photos SET is_current=0 WHERE accused_id=%s", (accused_id,))
            cursor.execute("""
                INSERT INTO accused_photos (accused_id, photo_url, photo_public_id, is_current, uploaded_by)
                VALUES (%s,%s,%s,1,%s)
            """, (accused_id, bail_photo_url, bail_photo_public_id, _uid()))

        cursor.execute("""
            UPDATE accused SET bail_status=%s, bail_start_date=%s, bail_end_date=%s,
            bail_documents_url=%s, bail_documents_public_id=%s,
            bail_photo_url=%s, bail_photo_public_id=%s,
            bail_photo_lat=%s, bail_photo_lng=%s, bail_photo_captured_at=%s,
            bail_remark=%s, bail_rating=%s, updated_by=%s
            {photo_cols}
            WHERE id=%s
        """.format(photo_cols=", photo_url=%s, photo_public_id=%s, profile_status='complete'" if set_as_profile_photo else ""),
        (bail_type, bail_start, bail_end, doc_url, doc_pub_id,
         bail_photo_url, bail_photo_public_id, photo_lat_val, photo_lng_val, photo_time_val,
         bail_remark, bail_rating, _uid(),
         *([bail_photo_url, bail_photo_public_id] if set_as_profile_photo else []),
         accused_id))

        cursor.execute("""
            INSERT INTO accused_bail_history
            (accused_id, fir_id, bail_type, bail_start_date, bail_end_date,
             bail_document_url, bail_document_public_id, bail_document_resource_type,
             bail_photo_url, bail_photo_public_id, bail_photo_lat, bail_photo_lng, bail_photo_captured_at,
             bail_remark, bail_rating, approved_by, status)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'ACTIVE')
        """, (accused_id, fir_id, bail_type, bail_start, bail_end,
              doc_url, doc_pub_id, doc_res_type or 'raw',
              bail_photo_url, bail_photo_public_id, photo_lat_val, photo_lng_val, photo_time_val,
              bail_remark, bail_rating, _uid()))
        bail_id = cursor.lastrowid

        # हर चुनी गई FIR को accused_bail_fir में जोड़ें (fir_id सहित, ताकि
        # junction table हमेशा पूर्ण रहे और डिस्प्ले क्वेरी उसी पर निर्भर हो सके)
        for fid in fir_ids:
            cursor.execute("""
                INSERT IGNORE INTO accused_bail_fir (bail_id, fir_id) VALUES (%s, %s)
            """, (bail_id, fid))
        conn.commit()

        fir_rows_selected = [f for f in arrest_firs if f['id'] in fir_ids]
        fir_label = ', '.join(f"{f['fir_number']}/{f['thana']}" for f in fir_rows_selected) or '—'
        fir_row = fir_rows_selected[0] if fir_rows_selected else None

        log_activity(_uid(), _role(),
                     f"Approved {bail_type} bail for accused ID:{accused_id} (FIR {fir_label})",
                     ip=request.remote_addr)

        try:
            from run import mail
        except Exception:
            mail = None
        send_bail_notification(
            district=fir_row['district'] if fir_row else _district(),
            accused_name=accused['name'],
            fir_label=f"S/o {accused['fathers_name']} | FIR {fir_label}",
            bail_type=bail_type, bail_start=bail_start, bail_end=bail_end,
            bail_remark=bail_remark, bail_rating=bail_rating,
            approved_by_name=session.get('name'), approved_by_id=_uid(),
            mail_instance=mail,
            thana=fir_row['thana'] if fir_row else None,
        )

        success_msg = 'जमानत सफलतापूर्वक स्वीकृत। जिले के अधिकारियों को सूचित किया गया।'
        if set_as_profile_photo:
            success_msg += ' चूंकि कोई प्रोफ़ाइल फ़ोटो नहीं थी, जमानत के समय की फ़ोटो को प्रोफ़ाइल फ़ोटो के रूप में सेट कर दिया गया है।'
        flash(success_msg, 'success')
        cursor.close(); conn.close()
        return redirect(url_for(f'{bp}.accused_detail', accused_id=accused_id))

    cursor.close(); conn.close()
    return render_template(f'{bp}/approve_accused_bail.html', accused=accused, arrest_firs=arrest_firs)


def revoke_accused_bail(accused_id: int, role='admin'):
    """POST-only: revoke the current active bail. History row stays (status=REVOKED)."""
    bp = 'admin' if role == 'admin' else 'super'
    revoke_reason = request.form.get('revoke_reason', '').strip()
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        UPDATE accused_bail_history
        SET status='REVOKED', revoked_by=%s, revoked_at=NOW(), revoke_reason=%s
        WHERE accused_id=%s AND status='ACTIVE'
        ORDER BY approved_at DESC LIMIT 1
    """, (_uid(), revoke_reason or None, accused_id))
    cursor.execute("""
        UPDATE accused SET bail_status='none', bail_start_date=NULL, bail_end_date=NULL,
        bail_documents_url=NULL, bail_documents_public_id=NULL,
        bail_photo_url=NULL, bail_photo_public_id=NULL,
        bail_photo_lat=NULL, bail_photo_lng=NULL, bail_photo_captured_at=NULL,
        bail_remark=NULL, bail_rating=0, updated_by=%s WHERE id=%s
    """, (_uid(), accused_id))
    conn.commit()
    log_activity(_uid(), _role(), f"Revoked bail for accused ID:{accused_id}", ip=request.remote_addr)
    flash('जमानत रद्द कर दी गई। इतिहास सुरक्षित रखा गया है।', 'success')
    cursor.close(); conn.close()
    return redirect(url_for(f'{bp}.accused_detail', accused_id=accused_id))


def get_bailed_accused_list(role='admin'):
    """List all accused with an approved bail (current or historical), with filters."""
    bp = 'admin' if role == 'admin' else 'super'
    district         = _district()
    status_filter    = request.args.get('status', 'ACTIVE')
    bail_type_filter = request.args.get('bail_type', '')
    search           = request.args.get('search', '').strip()
    page             = int(request.args.get('page', 1))
    per_page         = int(request.args.get('per_page', 25))

    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    conditions = ["f.district=%s"]
    params = [district]
    if status_filter in ('ACTIVE', 'REVOKED', 'COMPLETED'):
        conditions.append("abh.status=%s"); params.append(status_filter)
    if bail_type_filter in ('temporary', 'permanent'):
        conditions.append("abh.bail_type=%s"); params.append(bail_type_filter)
    if search:
        conditions.append("(a.name LIKE %s OR a.fathers_name LIKE %s)")
        like = f'%{search}%'; params += [like, like]

    where = " AND ".join(conditions)
    base_q = f"""
        SELECT abh.id AS bail_id, abh.bail_type, abh.bail_start_date, abh.bail_end_date,
               abh.bail_remark, abh.bail_rating, abh.status AS bail_history_status,
               abh.bail_document_url, abh.approved_at, abh.revoked_at, abh.completed_at,
               abh.revoke_reason, abh.photo_status, abh.source,
               a.id, a.name, a.fathers_name, a.photo_url,
               f.fir_number, f.thana, f.district,
               GROUP_CONCAT(DISTINCT CONCAT(f2.fir_number,'/',f2.thana)
                            ORDER BY f2.fir_number SEPARATOR ', ') AS all_firs,
               g.name AS approved_by_name, r.name AS revoked_by_name
        FROM accused_bail_history abh
        JOIN accused a ON a.id = abh.accused_id
        JOIN fir_cases f ON f.id = abh.fir_id
        LEFT JOIN accused_bail_fir abf ON abf.bail_id = abh.id
        LEFT JOIN fir_cases f2 ON f2.id = abf.fir_id
        LEFT JOIN users g ON g.id = abh.approved_by
        LEFT JOIN users r ON r.id = abh.revoked_by
        WHERE {where}
        GROUP BY abh.id
        ORDER BY abh.approved_at DESC
    """
    rows, total, total_pages = paginate_query(cursor, base_q, params, page, per_page)
    for b in rows:
        if not b.get('all_firs'):
            b['all_firs'] = f"{b['fir_number']}/{b['thana']}"
    cursor.close(); conn.close()

    tmpl = f'{bp}/bailed_accused.html'
    return render_template(
        tmpl, bail_records=rows, status_filter=status_filter,
        bail_type_filter=bail_type_filter, search=search,
        page=page, total=total, total_pages=total_pages, per_page=per_page
    )


# ── Manual FIR Case Creation ──────────────────────────────────────────────────

def create_fir_manual(role='admin'):
    """
    Manually create a FIR case with accused list entered as text.
    The accused text is parsed using the same s/o logic as Excel upload.
    """
    tmpl = f'{_bp()}/add_fir.html'

    if request.method == 'POST':
        district    = _district()
        thana       = request.form.get('thana', '').strip()
        fir_number  = request.form.get('fir_number', '').strip()
        acts        = request.form.get('acts', '').strip()
        total_raw   = request.form.get('total_accused', '').strip()
        fir_raw     = request.form.get('fir_accused', '').strip()
        arrested_raw= request.form.get('arrested_accused', '').strip()
        cs_raw      = request.form.get('cs_accused', '').strip()
        complainant = request.form.get('complainant', '').strip()
        status      = request.form.get('status', '').strip()

        if not thana or not fir_number:
            flash('थाना और FIR संख्या आवश्यक है।', 'danger')
            return render_template(tmpl)

        conn   = get_connection()
        cursor = conn.cursor(dictionary=True)

        try:
            cursor.execute("""
                INSERT INTO fir_cases
                    (district, thana, fir_number, acts,
                     total_accused_raw, fir_accused_raw,
                     arrested_accused_raw, cs_accused_raw,
                     complainant, status, created_by)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON DUPLICATE KEY UPDATE
                    acts=VALUES(acts), complainant=VALUES(complainant),
                    status=VALUES(status),
                    total_accused_raw=VALUES(total_accused_raw),
                    fir_accused_raw=VALUES(fir_accused_raw),
                    arrested_accused_raw=VALUES(arrested_accused_raw),
                    cs_accused_raw=VALUES(cs_accused_raw)
            """, (district, thana, fir_number, acts,
                  total_raw, fir_raw, arrested_raw, cs_raw,
                  complainant, status, _uid()))

            cursor.execute(
                "SELECT id FROM fir_cases WHERE district=%s AND thana=%s AND fir_number=%s",
                (district, thana, fir_number)
            )
            fir_id = cursor.fetchone()['id']

            total_set    = set(normalize_name(a['name']) for a in parse_accused_list(total_raw))
            fir_set      = set(normalize_name(a['name']) for a in parse_accused_list(fir_raw))
            arrested_set = set(normalize_name(a['name']) for a in parse_accused_list(arrested_raw))
            cs_set       = set(normalize_name(a['name']) for a in parse_accused_list(cs_raw))

            # parse_accused_list() already drops blank/अज्ञात/unknown names
            all_accused = parse_accused_list(total_raw)
            seen_norms  = set(total_set)
            for a in parse_accused_list(fir_raw):
                n = normalize_name(a['name'])
                if n not in seen_norms:
                    all_accused.append(a)
                    seen_norms.add(n)

            linked_count = 0
            for acc in all_accused:
                acc_id, _created = find_or_create_accused(cursor, acc['name'], acc['fathers_name'], _uid())
                if acc_id is None:
                    continue
                n = normalize_name(acc['name'])
                upsert_accused_fir(
                    cursor, acc_id, fir_id,
                    n in total_set, n in fir_set,
                    n in arrested_set, n in cs_set
                )
                linked_count += 1

            cursor.execute(
                "UPDATE fir_cases SET total_accused_count=%s WHERE id=%s",
                (linked_count, fir_id)
            )
            conn.commit()
            log_activity(_uid(), _role(), f"FIR {fir_number}/{thana} manually created",
                         ip=request.remote_addr)
            flash(f'FIR {fir_number} सफलतापूर्वक दर्ज किया गया। {linked_count} अभियुक्त जोड़े गए।',
                  'success')
            cursor.close(); conn.close()
            return redirect(url_for(f'{_bp()}.fir_detail', fir_id=fir_id))

        except Exception as e:
            conn.rollback()
            logger.error(f"Manual FIR creation error: {e}")
            flash(f'त्रुटि: {e}', 'danger')
            cursor.close(); conn.close()
            return render_template(tmpl)

    return render_template(tmpl)


# ── Sample Excel Download ─────────────────────────────────────────────────────

def download_accused_sample_file():
    """Return a sample Excel in the correct UP Police format."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'आपराधिक प्रकरण'

        headers = [
            'क्र.', 'जनपद', 'थाना', 'FIR संख्या', 'धारा',
            'कुल अभियुक्त', 'FIR में अभियुक्त',
            'गिरफ्तार अभियुक्त', 'आरोप पत्र अभियुक्त',
            'वादी', 'स्थिति'
        ]
        ws.append(headers)

        # Style header
        header_fill = PatternFill("solid", fgColor="1a73e8")
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF", size=11)
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        # Sample row
        ws.append([
            1, 'मिर्जापुर', 'जमालपुर', '0001/2025',
            'भारतीय न्याय संहिता 2023 धारा 303, 351',
            'रामकुमार s/o श्यामलाल, मोहन सिंह s/o रामसिंह',
            'रामकुमार s/o श्यामलाल',
            'रामकुमार s/o श्यामलाल',
            'रामकुमार s/o श्यामलाल',
            'विनोद यादव',
            'CS-01/06/2025'
        ])

        # Column widths
        widths = [5, 15, 15, 15, 35, 45, 45, 45, 45, 20, 15]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64+i)].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True, download_name='accused_upload_sample.xlsx'
        )
    except ImportError:
        flash('openpyxl install करें: pip install openpyxl', 'danger')
        return redirect(url_for(f'{_bp()}.upload_accused'))