"""
bail_bulk.py
============
जमानत (Bail) से जुड़े दो नए features:

  1) एक जमानत रिकॉर्ड में एक से अधिक FIR चुनने का समर्थन (multi-FIR bail),
     ताकि एक ही अभियुक्त की कई FIR एक साथ एक जमानत में कवर हो सकें।

  2) माननीय न्यायालय द्वारा जमानत/रिहाई की Excel सूची अपलोड कर, अनेक
     अभियुक्तों की जमानत एक साथ स्वीकृत करना — नाम व पिता के नाम पर
     fuzzy-matching करके, थाना/FIR संख्या से क्रॉस-चेक करके।

     प्रवाह (2-चरण, ताकि कोई कानूनी रिकॉर्ड गलत मैच से न बन जाए):
       a) upload  -> parse हर पंक्ति -> हर पंक्ति का सर्वश्रेष्ठ मैच खोजें
                     -> bail_excel_batch + bail_excel_row में "staged" सेव करें
       b) review  -> admin को matched / ambiguous / not_found / already_bailed
                     / fir_not_found — सभी श्रेणियाँ साफ़-साफ़ दिखाई जाती हैं
       c) confirm -> admin जिन पंक्तियों को स्वीकार करता है, केवल उन्हीं के
                     लिए वास्तविक accused_bail_history रिकॉर्ड बनते हैं;
                     फ़ोटो/दस्तावेज़ photo_status='pending' रहते हैं और बाद
                     में "लंबित फ़ोटो" सूची से अलग से अपलोड किए जा सकते हैं।

Not-found / already-approved accused for a row are NEVER silently dropped —
they always surface in the batch's error report (see get_batch_review()).
"""

import re
import json
import logging
from datetime import datetime, date

from flask import session, request, redirect, url_for, flash, render_template

from db import get_connection
from utils import upload_image, upload_document, log_activity

logger = logging.getLogger(__name__)


# ── Fuzzy-match engine ──────────────────────────────────────────────────────
# rapidfuzz (C-accelerated) is used when available — far faster than
# difflib.SequenceMatcher at production batch sizes (hundreds of Excel rows
# x thousands of candidate accused per district), and more forgiving of the
# small typos/transpositions common in hand-typed Hindi court sheets. Falls
# back to the stdlib difflib ratio if rapidfuzz isn't installed, so this
# module never hard-fails on a missing optional dependency.
try:
    from rapidfuzz import fuzz as _rf_fuzz
    from rapidfuzz.distance import JaroWinkler as _rf_jw

    def _raw_similarity(a: str, b: str) -> float:
        # Blend token-sort ratio (robust to word-order swaps, e.g. someone
        # typing "सोनकर करन" instead of "करन सोनकर") with Jaro-Winkler
        # (rewards long common prefixes, typical of Indian given names).
        ts = _rf_fuzz.token_sort_ratio(a, b) / 100.0
        jw = _rf_jw.similarity(a, b)
        return max(ts, jw)

    _FUZZY_ENGINE = 'rapidfuzz'
except ImportError:  # pragma: no cover - exercised only when dep missing
    from difflib import SequenceMatcher

    def _raw_similarity(a: str, b: str) -> float:
        return SequenceMatcher(None, a, b).ratio()

    _FUZZY_ENGINE = 'difflib'

logger.info(f"bail_bulk fuzzy-match engine: {_FUZZY_ENGINE}")


# ── session helpers (mirrors accused_common.py) ────────────────────────────

def _district(): return session.get('district')
def _uid():      return session.get('user_id')
def _role():     return session.get('role')
def _bp():       return 'admin' if _role() == 'admin' else 'super'


# ── Name / address parsing for the "अभियुक्त का नाम पता" Excel column ──────
#
# Real-world format seen in court/jail release sheets:
#   "करन सोनकर पुत्र राजेश निवासी पाण्डेयपुर थाना को0 शहर मीरजापुर"
#    <name>       <पुत्र> <father>  <निवासी> <address ... थाना THANA DISTRICT>

_FATHER_SPLIT_RE = re.compile(
    r'\s+(?:s/o|S/O|पुत्र|पुत्री|सुपुत्र|सुपुत्री|आत्मज|आत्मजा|w/o|W/O|d/o|D/O|पति|पत्नी)\s+',
    re.IGNORECASE
)
_ADDRESS_SPLIT_RE = re.compile(r'\s+(?:निवासी|नि0|साकिन)\s+')
_THANA_IN_ADDRESS_RE = re.compile(r'थाना\s+(.+)$')
_PUNCT_CLEAN_RE = re.compile(r'[.,;:।]+')


def parse_bail_name_address(raw: str) -> dict:
    """
    Parse one "अभियुक्त का नाम पता" cell into name / father's name / address.
    Tolerant of missing pieces — always returns a dict with all four keys,
    falling back to empty strings rather than raising.
    """
    raw = re.sub(r'\s+', ' ', (raw or '')).strip()
    if not raw:
        return {'name': '', 'fathers_name': '', 'address': '', 'thana_guess': ''}

    addr_parts = _ADDRESS_SPLIT_RE.split(raw, maxsplit=1)
    name_father_part = addr_parts[0].strip()
    address = addr_parts[1].strip() if len(addr_parts) == 2 else ''

    m = _FATHER_SPLIT_RE.split(name_father_part, maxsplit=1)
    if len(m) == 2:
        name, fathers_name = m[0].strip(), m[1].strip()
    else:
        name, fathers_name = name_father_part.strip(), ''

    thana_guess = ''
    if address:
        tm = _THANA_IN_ADDRESS_RE.search(address)
        if tm:
            thana_guess = tm.group(1).strip()

    return {
        'name': name,
        'fathers_name': fathers_name,
        'address': address,
        'thana_guess': thana_guess,
    }


_NAME_PREFIXES = ['श्री ', 'श्रीमती ', 'सुश्री ', 'स्व0 ', 'स्व ', 'स्वर्गीय ',
                  'Mr. ', 'Mrs. ', 'Smt. ', 'Shri ', 'Ms. ', 'Dr. ']


def normalize_name(name: str) -> str:
    """Same normalization as accused_common.normalize_name (kept local to
    avoid a circular import — both modules must stay in sync).
    Strips honorifics, punctuation, and collapses whitespace so trivial
    formatting differences ("श्री करन सोनकर" vs "करन सोनकर," vs "करन  सोनकर")
    never cost match score."""
    if not name:
        return ''
    n = name.strip()
    changed = True
    while changed:
        changed = False
        for prefix in _NAME_PREFIXES:
            if n.lower().startswith(prefix.lower()):
                n = n[len(prefix):].strip()
                changed = True
    n = _PUNCT_CLEAN_RE.sub(' ', n)
    return re.sub(r'\s+', ' ', n).strip().lower()


def normalize_thana(thana: str) -> str:
    """Normalize a थाना string for loose comparison: strip a leading
    'थाना' word, drop stray '0' digits (a common OCR/typing artifact for
    '।'/full-stops in UP Police sheets, e.g. 'को0 देहात' / 'को0शहर'),
    strip punctuation, and collapse whitespace."""
    if not thana:
        return ''
    t = thana.strip()
    t = re.sub(r'^थाना\s*', '', t)
    t = t.replace('0', '')             # 'को0देहात' -> 'कोदेहात'
    t = _PUNCT_CLEAN_RE.sub(' ', t)
    t = re.sub(r'\s+', ' ', t).strip().lower()
    return t


def name_similarity(a: str, b: str) -> float:
    """0..1 similarity between two normalized strings, via the configured
    fuzzy engine (rapidfuzz if installed, else difflib)."""
    if not a or not b:
        return 0.0
    return _raw_similarity(a, b)


def weighted_match_score(name_a, name_b, father_a, father_b) -> tuple:
    """
    Returns (combined_score, name_score, father_score).
    Name carries more weight (0.6) since it's the primary identifier;
    father's name (0.4) disambiguates common first names — very common
    in Hindi naming, e.g. many "राम कुमार" but few share both name AND
    father's name.
    """
    ns = name_similarity(name_a, name_b)
    fs = name_similarity(father_a, father_b)
    return (0.6 * ns + 0.4 * fs), ns, fs


MATCH_THRESHOLD        = 0.80   # combined score to count as a real candidate
MIN_NAME_THRESHOLD     = 0.72   # name alone must be at least this close
AMBIGUOUS_GAP          = 0.04   # if top-2 candidates are within this gap -> ambiguous


# ── Date parsing (Excel cells arrive as datetime, date, or DD.MM.YYYY text) ─

def parse_excel_date(val):
    if val is None or val == '':
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ('%d.%m.%Y', '%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


# ── Candidate lookup ─────────────────────────────────────────────────────────

def find_accused_candidates(cursor, district: str, name: str, fathers_name: str,
                             thana_col: str = '', fir_number_col: str = '', limit: int = 5):
    """
    Find the best-matching *bail-eligible* accused (arrested in >=1 FIR of
    this district, and currently has no active bail) for a parsed
    name/father pair from an uploaded bail Excel row.

    Scoping strategy (most reliable first):
      1. If both थाना and FIR संख्या columns resolve to a known fir_cases
         row, restrict candidates to accused linked to that exact FIR.
      2. Else if थाना alone resolves, restrict to accused arrested in that
         थाना within the district.
      3. Else fall back to every bail-eligible accused in the district.

    Returns (candidates, resolved_fir_id, fir_lookup_status) where
    candidates is a list of dicts sorted by score desc, each with:
      accused_id, name, fathers_name, score, name_score, father_score,
      fir_ids (list of int — this accused's arrest FIRs in the district).
    fir_lookup_status is one of 'exact', 'thana_only', 'none'.
    """
    name_norm   = normalize_name(name)
    father_norm = normalize_name(fathers_name)
    thana_norm  = normalize_thana(thana_col)
    fir_number  = (fir_number_col or '').strip()

    resolved_fir_id  = None
    fir_lookup_status = 'none'

    if thana_norm and fir_number:
        cursor.execute("""
            SELECT id, thana FROM fir_cases
            WHERE district=%s AND fir_number=%s
        """, (district, fir_number))
        for row in cursor.fetchall():
            if normalize_thana(row['thana']) == thana_norm:
                resolved_fir_id = row['id']
                fir_lookup_status = 'exact'
                break

    # Base pool: bail-eligible accused (arrested, no active bail) in district
    base_sql = """
        SELECT DISTINCT a.id, a.name, a.fathers_name, a.name_normalized, a.fathers_normalized
        FROM accused a
        JOIN accused_fir af ON af.accused_id = a.id
        JOIN fir_cases f ON f.id = af.fir_id
        WHERE f.district = %s AND af.in_arrested = 1
        AND (a.bail_status IS NULL OR a.bail_status = 'none')
    """
    params = [district]
    if resolved_fir_id:
        base_sql += " AND af.fir_id = %s"
        params.append(resolved_fir_id)
    elif thana_norm:
        base_sql += " AND f.thana = %s"
        params.append(thana_col.strip())
        fir_lookup_status = 'thana_only'

    cursor.execute(base_sql, params)
    pool = cursor.fetchall()

    # If the thana-scoped pool is empty (e.g. thana text mismatch/typo),
    # fall back to the whole district rather than reporting a false negative.
    if not pool and (resolved_fir_id or thana_norm):
        cursor.execute("""
            SELECT DISTINCT a.id, a.name, a.fathers_name, a.name_normalized, a.fathers_normalized
            FROM accused a
            JOIN accused_fir af ON af.accused_id = a.id
            JOIN fir_cases f ON f.id = af.fir_id
            WHERE f.district = %s AND af.in_arrested = 1
            AND (a.bail_status IS NULL OR a.bail_status = 'none')
        """, (district,))
        pool = cursor.fetchall()
        fir_lookup_status = 'none'
        resolved_fir_id = None

    scored = []
    for cand in pool:
        score, ns, fs = weighted_match_score(
            name_norm, cand['name_normalized'], father_norm, cand['fathers_normalized']
        )
        if score >= MATCH_THRESHOLD and ns >= MIN_NAME_THRESHOLD:
            scored.append({
                'accused_id': cand['id'], 'name': cand['name'],
                'fathers_name': cand['fathers_name'],
                'score': round(score, 4), 'name_score': round(ns, 4),
                'father_score': round(fs, 4),
            })

    scored.sort(key=lambda c: c['score'], reverse=True)
    scored = scored[:limit]

    # Attach each candidate's arrest-FIRs in this district (for display +
    # so confirm() knows which FIR(s) to attach the new bail record to).
    for c in scored:
        cursor.execute("""
            SELECT DISTINCT f.id FROM accused_fir af
            JOIN fir_cases f ON f.id = af.fir_id
            WHERE af.accused_id=%s AND af.in_arrested=1 AND f.district=%s
        """, (c['accused_id'], district))
        c['fir_ids'] = [r['id'] for r in cursor.fetchall()]

    return scored, resolved_fir_id, fir_lookup_status


def classify_match(candidates, resolved_fir_id):
    """
    Decide match_status from a scored candidate list.
    Returns (status, matched_accused_id_or_None, resolved_fir_ids_list).
    """
    if not candidates:
        return 'not_found', None, []

    if len(candidates) >= 2 and (candidates[0]['score'] - candidates[1]['score']) < AMBIGUOUS_GAP:
        return 'ambiguous', None, []

    best = candidates[0]
    fir_ids = [resolved_fir_id] if resolved_fir_id else best['fir_ids']
    return 'matched', best['accused_id'], fir_ids


# ── Stage: parse the uploaded Excel + fuzzy-match every row ─────────────────

# Expected header (UP Police jail-release format), 0-indexed column order:
#  0 क्र0सं0 | 1 थाना | 2 मु0अ0सं0 | 3 एसटी नं0 | 4 भादवि0 की धारा |
#  5 बीएनएस की धारा | 6 अभियुक्त का नाम पता | 7 अभियुक्त का फोटो |
#  8 अपराधिक इतिहास | 9 मा0 न्यायालय का नाम | 10 जमानत दिये जाने का दिनांक |
#  11 जेल जाने का दिनांक | 12 रिहा किये जाने की तिथी | 13 चौकी/हल्का | 14 बीट सं0
BAIL_EXCEL_COLUMNS = 15

# Production guardrail: a single court-order sheet is a few hundred rows at
# most in real usage; capping this protects the request/DB from an
# accidental (or malicious) multi-thousand-row upload tying up a worker.
MAX_BAIL_EXCEL_ROWS = 3000

# Header detection: prior versions matched any row containing the substring
# 'अभियुक्त' anywhere — which false-positives on a *title* row like
# "...जेल से रिहा हुए अभियुक्तगण का विवरण दिनांक..." (a very common first
# line in real UP-Police jail-release sheets, incl. the one this system
# ships a sample of). That caused the real header row to be read as data.
# Strong markers are phrases that only ever appear in the actual column
# header; weak markers alone must co-occur (>=2) to count, so a title
# mentioning one police/legal term in passing doesn't get mistaken for it.
_HEADER_STRONG_MARKERS = ['नाम पता', 'नाम  पता', 'नाम एवं पता']
_HEADER_WEAK_MARKERS = ['क्र0सं0', 'क्र0', 'थाना', 'मु0अ0सं0', 'अभियुक्त', 'फोटो', 'न्यायालय']


def _looks_like_header_row(row) -> bool:
    cells = [str(c) for c in row if c not in (None, '')]
    if not cells:
        return False
    text = ' '.join(cells)
    if any(marker in text for marker in _HEADER_STRONG_MARKERS):
        return True
    hits = sum(1 for marker in _HEADER_WEAK_MARKERS if marker in text)
    return hits >= 2


def stage_bail_excel(file, filename: str, district: str, uploaded_by: int) -> int:
    """
    Parse + fuzzy-match every accused row in the uploaded Excel, persist
    the batch + all rows (matched and unmatched alike) to the database,
    and return the new batch_id. Does NOT create any bail record yet —
    that only happens on confirm_batch().

    Raises ValueError with a Hindi, user-facing message on any parse
    failure (bad file, too many rows, no header found, etc); the staged
    batch row (if already created) is cleaned up before re-raising so a
    failed upload never leaves a phantom empty batch behind.
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(file, data_only=True, read_only=True)
        ws = wb.active
    except Exception as e:
        logger.error(f"stage_bail_excel: workbook load failed: {e}")
        raise ValueError('Excel फ़ाइल पढ़ी नहीं जा सकी। कृपया मान्य .xlsx फ़ाइल अपलोड करें।')

    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    batch_id = None

    try:
        cursor.execute("""
            INSERT INTO bail_excel_batch (district, filename, uploaded_by, status)
            VALUES (%s, %s, %s, 'staged')
        """, (district, filename, uploaded_by))
        batch_id = cursor.lastrowid

        # Find the real header row by scanning further than before (title
        # blocks can span 1-2 merged rows) and using the stronger detector.
        header_row_idx = None
        scan_limit = min(10, ws.max_row or 10)
        for idx, row in enumerate(
                ws.iter_rows(min_row=1, max_row=scan_limit, values_only=True), start=1):
            if row and _looks_like_header_row(row):
                header_row_idx = idx
                break
        if header_row_idx is None:
            # Nothing matched the header heuristic at all — safest fallback
            # is row 1 (better to mis-parse one junk row than silently skip
            # every real data row), but log loudly so it's caught in review.
            header_row_idx = 1
            logger.warning(
                f"stage_bail_excel: no header row detected in '{filename}', "
                f"defaulting to row 1 — verify column mapping in review."
            )

        total_rows, matched_rows, error_rows = 0, 0, 0

        for row_idx, row in enumerate(
                ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
            if not row or not any(row):
                continue
            if total_rows >= MAX_BAIL_EXCEL_ROWS:
                logger.warning(
                    f"stage_bail_excel: row cap ({MAX_BAIL_EXCEL_ROWS}) hit for "
                    f"batch {batch_id}, remaining rows skipped."
                )
                break
            cells = (list(row) + [None] * BAIL_EXCEL_COLUMNS)[:BAIL_EXCEL_COLUMNS]
            (_sr, thana_col, fir_number_col, _st_no, _ipc, _bns,
             accused_field, _photo, criminal_history, court_name,
             bail_date_raw, jail_date_raw, release_date_raw, _chowki, _beat) = cells

            accused_field = str(accused_field or '').strip()
            if not accused_field:
                continue  # blank accused cell — nothing to import for this row

            total_rows += 1
            parsed = parse_bail_name_address(accused_field)
            thana_col_s = str(thana_col or '').strip()
            fir_number_s = str(fir_number_col or '').strip()

            if not parsed['name']:
                cursor.execute("""
                    INSERT INTO bail_excel_row
                        (batch_id, row_no, raw_name_field, parsed_name, parsed_fathers_name,
                         parsed_address, thana_col, fir_number_col, court_name, jail_date,
                         release_date, bail_date, criminal_history, match_status, include_in_confirm)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'not_found',0)
                """, (batch_id, row_idx, accused_field, '', '', parsed['address'],
                      thana_col_s, fir_number_s, str(court_name or '').strip(),
                      parse_excel_date(jail_date_raw), parse_excel_date(release_date_raw),
                      parse_excel_date(bail_date_raw), str(criminal_history or '').strip()))
                error_rows += 1
                continue

            candidates, resolved_fir_id, _lookup = find_accused_candidates(
                cursor, district, parsed['name'], parsed['fathers_name'],
                thana_col_s, fir_number_s
            )
            status, matched_id, fir_ids = classify_match(candidates, resolved_fir_id)

            # already_bailed check: only relevant if the *exact* normalized
            # name+father exists but is currently excluded from the eligible
            # pool because they already have an active bail.
            if status == 'not_found':
                name_norm = normalize_name(parsed['name'])
                father_norm = normalize_name(parsed['fathers_name'])
                cursor.execute("""
                    SELECT a.id FROM accused a
                    WHERE a.name_normalized = %s AND a.fathers_normalized = %s
                    AND a.bail_status != 'none'
                    LIMIT 1
                """, (name_norm, father_norm))
                if cursor.fetchone():
                    status = 'already_bailed'

            if status == 'matched':
                matched_rows += 1
            else:
                error_rows += 1

            cursor.execute("""
                INSERT INTO bail_excel_row
                    (batch_id, row_no, raw_name_field, parsed_name, parsed_fathers_name,
                     parsed_address, thana_col, fir_number_col, court_name, jail_date,
                     release_date, bail_date, criminal_history, match_status,
                     matched_accused_id, match_confidence, candidate_ids, resolved_fir_ids,
                     include_in_confirm)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                batch_id, row_idx, accused_field, parsed['name'], parsed['fathers_name'],
                parsed['address'], thana_col_s, fir_number_s, str(court_name or '').strip(),
                parse_excel_date(jail_date_raw), parse_excel_date(release_date_raw),
                parse_excel_date(bail_date_raw), str(criminal_history or '').strip(),
                status, matched_id,
                candidates[0]['score'] if candidates else None,
                ','.join(str(c['accused_id']) for c in candidates) if status == 'ambiguous' else None,
                ','.join(str(i) for i in fir_ids) if fir_ids else None,
                1 if status == 'matched' else 0,
            ))

        if total_rows == 0:
            raise ValueError(
                'Excel में कोई मान्य अभियुक्त पंक्ति नहीं मिली। कृपया जाँचें कि "अभियुक्त '
                'का नाम पता" कॉलम भरा है और फ़ाइल सही टेम्पलेट में है।'
            )

        cursor.execute("""
            UPDATE bail_excel_batch SET total_rows=%s, matched_rows=%s, error_rows=%s WHERE id=%s
        """, (total_rows, matched_rows, error_rows, batch_id))
        conn.commit()
        return batch_id

    except ValueError:
        conn.rollback()
        if batch_id:
            try:
                cursor.execute("DELETE FROM bail_excel_batch WHERE id=%s", (batch_id,))
                conn.commit()
            except Exception:
                pass
        raise
    except Exception as e:
        logger.error(f"stage_bail_excel: unexpected error for '{filename}': {e}")
        conn.rollback()
        if batch_id:
            try:
                cursor.execute("DELETE FROM bail_excel_batch WHERE id=%s", (batch_id,))
                conn.commit()
            except Exception:
                pass
        raise ValueError('Excel प्रोसेस करते समय त्रुटि हुई। फ़ाइल फ़ॉर्मैट जाँचें और पुनः प्रयास करें।')
    finally:
        cursor.close()
        conn.close()


# ── Review: fetch a staged batch for admin confirmation ─────────────────────

MATCH_STATUS_LABELS = {
    'matched':        'मैच मिला',
    'ambiguous':      'एक से अधिक संभावित मैच — मैन्युअल जाँच आवश्यक',
    'not_found':      'सिस्टम में यह अभियुक्त नहीं मिला',
    'already_bailed': 'इस अभियुक्त की जमानत पहले से स्वीकृत है',
    'fir_not_found':  'FIR सिस्टम में नहीं मिली',
}


def get_batch_review(batch_id: int, district: str):
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM bail_excel_batch WHERE id=%s AND district=%s", (batch_id, district))
    batch = cursor.fetchone()
    if not batch:
        cursor.close(); conn.close()
        return None, []

    cursor.execute("""
        SELECT r.*, a.name AS matched_name, a.fathers_name AS matched_fathers_name,
               a.photo_url AS matched_photo_url
        FROM bail_excel_row r
        LEFT JOIN accused a ON a.id = r.matched_accused_id
        WHERE r.batch_id=%s
        ORDER BY r.row_no
    """, (batch_id,))
    rows = cursor.fetchall()

    for r in rows:
        r['status_label'] = MATCH_STATUS_LABELS.get(r['match_status'], r['match_status'])
        if r['match_status'] == 'ambiguous' and r['candidate_ids']:
            ids = [int(x) for x in r['candidate_ids'].split(',') if x]
            if ids:
                fmt_ids = ','.join(['%s'] * len(ids))
                cursor.execute(f"SELECT id, name, fathers_name FROM accused WHERE id IN ({fmt_ids})", ids)
                r['candidates'] = cursor.fetchall()
            else:
                r['candidates'] = []
        else:
            r['candidates'] = []
        if r['resolved_fir_ids']:
            fids = [int(x) for x in r['resolved_fir_ids'].split(',') if x]
            if fids:
                fmt_ids = ','.join(['%s'] * len(fids))
                cursor.execute(
                    f"SELECT id, fir_number, thana FROM fir_cases WHERE id IN ({fmt_ids})", fids)
                r['resolved_firs'] = cursor.fetchall()
            else:
                r['resolved_firs'] = []
        else:
            r['resolved_firs'] = []

    cursor.close()
    conn.close()
    return batch, rows


# ── Resolve an 'ambiguous' row: admin manually picks the correct candidate ──

def resolve_ambiguous_row(batch_id: int, row_id: int, accused_id: int, district: str):
    """
    Admin looked at an 'ambiguous' row's candidate list (2+ near-equal
    fuzzy matches) and picked the correct accused by eye. Promotes the row
    to 'matched' against that accused_id so it becomes eligible for
    confirm_batch(). Refuses silently-wrong picks by validating the
    accused_id was actually one of the offered candidates for this row.
    """
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT r.*, b.district FROM bail_excel_row r
        JOIN bail_excel_batch b ON b.id = r.batch_id
        WHERE r.id=%s AND r.batch_id=%s AND b.district=%s
    """, (row_id, batch_id, district))
    row = cursor.fetchone()
    if not row or row['match_status'] != 'ambiguous':
        cursor.close(); conn.close()
        return False, 'यह पंक्ति ambiguous स्थिति में नहीं है।'

    candidate_ids = [int(x) for x in (row['candidate_ids'] or '').split(',') if x]
    if accused_id not in candidate_ids:
        cursor.close(); conn.close()
        return False, 'चुना गया अभियुक्त इस पंक्ति के संभावित मैचों में नहीं है।'

    cursor.execute("""
        SELECT DISTINCT f.id FROM accused_fir af
        JOIN fir_cases f ON f.id = af.fir_id
        WHERE af.accused_id=%s AND af.in_arrested=1 AND f.district=%s
    """, (accused_id, district))
    fir_ids = [r['id'] for r in cursor.fetchall()]
    if not fir_ids:
        cursor.close(); conn.close()
        return False, 'चुने गए अभियुक्त की इस जिले में कोई गिरफ़्तारी FIR नहीं मिली।'

    cursor.execute("""
        UPDATE bail_excel_row
        SET match_status='matched', matched_accused_id=%s,
            resolved_fir_ids=%s, include_in_confirm=1
        WHERE id=%s
    """, (accused_id, ','.join(str(i) for i in fir_ids), row_id))
    cursor.execute("""
        UPDATE bail_excel_batch
        SET matched_rows = matched_rows + 1, error_rows = error_rows - 1
        WHERE id=%s
    """, (batch_id,))
    conn.commit()
    cursor.close(); conn.close()
    return True, 'पंक्ति सफलतापूर्वक मैच कर दी गई — अब confirm करने पर जमानत बनेगी।'


# ── Discard a staged batch without creating any bail record ─────────────────

def discard_batch(batch_id: int, district: str, uid: int):
    """Abandon a staged batch entirely (e.g. wrong file uploaded). Rows are
    cascade-deleted with the batch; no accused/bail data is ever touched
    since nothing is written outside bail_excel_* until confirm_batch()."""
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM bail_excel_batch WHERE id=%s AND district=%s", (batch_id, district))
    batch = cursor.fetchone()
    if not batch:
        cursor.close(); conn.close()
        return False, 'बैच नहीं मिला।'
    if batch['status'] != 'staged':
        cursor.close(); conn.close()
        return False, 'यह बैच पहले ही प्रोसेस/डिसकार्ड हो चुका है।'
    cursor.execute(
        "UPDATE bail_excel_batch SET status='discarded', confirmed_by=%s, confirmed_at=NOW() WHERE id=%s",
        (uid, batch_id)
    )
    conn.commit()
    cursor.close(); conn.close()
    return True, 'बैच डिसकार्ड कर दिया गया — कोई जमानत रिकॉर्ड नहीं बनाया गया।'


# ── Batch history (all uploads for this district, any status) ───────────────

def list_batches(district: str, limit: int = 50):
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT b.*, u.name AS uploaded_by_name, c.name AS confirmed_by_name
        FROM bail_excel_batch b
        LEFT JOIN users u ON u.id = b.uploaded_by
        LEFT JOIN users c ON c.id = b.confirmed_by
        WHERE b.district=%s
        ORDER BY b.created_at DESC
        LIMIT %s
    """, (district, limit))
    rows = cursor.fetchall()
    cursor.close(); conn.close()
    return rows


# ── Confirm: create real bail records for the rows admin approved ───────────

def confirm_batch(batch_id: int, district: str, confirmed_row_ids: list, resolved_by_uid: int):
    """
    For every bail_excel_row in this batch whose id is in confirmed_row_ids
    AND whose match_status == 'matched' (ambiguous/not_found/already_bailed
    rows are never auto-confirmed, even if somehow included), create an
    accused_bail_history record with photo_status='pending' (the photo /
    supporting document is uploaded afterwards from the "लंबित फ़ोटो" list),
    attach every resolved FIR via accused_bail_fir, and update the accused's
    current bail_status.

    Returns dict: {created: [...], skipped: [...]}
    """
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM bail_excel_batch WHERE id=%s AND district=%s", (batch_id, district))
    batch = cursor.fetchone()
    if not batch or batch['status'] != 'staged':
        cursor.close(); conn.close()
        return {'created': [], 'skipped': [], 'error': 'बैच पहले ही प्रोसेस हो चुका है या नहीं मिला।'}

    cursor.execute("SELECT * FROM bail_excel_row WHERE batch_id=%s", (batch_id,))
    all_rows = {r['id']: r for r in cursor.fetchall()}

    created, skipped = [], []

    for row_id in confirmed_row_ids:
        r = all_rows.get(row_id)
        if not r or r['match_status'] != 'matched' or not r['matched_accused_id']:
            if r:
                skipped.append({'row': r['row_no'], 'reason': 'मैच नहीं — auto-skip'})
            continue

        accused_id = r['matched_accused_id']
        cursor.execute("SELECT bail_status FROM accused WHERE id=%s", (accused_id,))
        acc = cursor.fetchone()
        if not acc or (acc['bail_status'] and acc['bail_status'] != 'none'):
            skipped.append({'row': r['row_no'], 'reason': 'अभियुक्त की जमानत पहले ही सक्रिय हो गई — दोबारा नहीं बनाई गई'})
            continue

        fir_ids = [int(x) for x in (r['resolved_fir_ids'] or '').split(',') if x]
        if not fir_ids:
            skipped.append({'row': r['row_no'], 'reason': 'कोई गिरफ़्तारी FIR नहीं मिली'})
            continue

        bail_type = 'temporary'
        bail_start = r['bail_date'] or date.today()
        bail_end = None  # court-order Excel imports don't carry an explicit end date

        cursor.execute("""
            INSERT INTO accused_bail_history
                (accused_id, fir_id, bail_type, bail_start_date, bail_end_date,
                 status, approved_by, court_name, jail_date, release_date,
                 criminal_history, photo_status, source, excel_batch_id, excel_row_raw)
            VALUES (%s,%s,%s,%s,%s,'ACTIVE',%s,%s,%s,%s,%s,'pending','excel_bulk',%s,%s)
        """, (
            accused_id, fir_ids[0], bail_type, bail_start, bail_end,
            resolved_by_uid, r['court_name'], r['jail_date'], r['release_date'],
            r['criminal_history'], batch_id,
            json.dumps({'raw_name_field': r['raw_name_field'], 'row_no': r['row_no']},
                       ensure_ascii=False),
        ))
        bail_id = cursor.lastrowid

        for fid in fir_ids:
            cursor.execute("""
                INSERT IGNORE INTO accused_bail_fir (bail_id, fir_id) VALUES (%s, %s)
            """, (bail_id, fid))

        cursor.execute("""
            UPDATE accused SET bail_status=%s, bail_start_date=%s, bail_end_date=%s,
            bail_remark=%s, updated_by=%s WHERE id=%s
        """, (bail_type, bail_start, bail_end,
              'Excel बल्क जमानत स्वीकृति — फ़ोटो/दस्तावेज़ लंबित', resolved_by_uid, accused_id))

        cursor.execute("UPDATE bail_excel_row SET created_bail_id=%s WHERE id=%s", (bail_id, r['id']))

        created.append({'row': r['row_no'], 'accused_id': accused_id,
                         'name': r['parsed_name'], 'bail_id': bail_id})

    cursor.execute("""
        UPDATE bail_excel_batch SET status='confirmed', confirmed_by=%s, confirmed_at=NOW() WHERE id=%s
    """, (resolved_by_uid, batch_id))
    conn.commit()
    cursor.close()
    conn.close()

    # ── Send FCM push + in-app notification for each approved bail ────────────
    # Done AFTER commit so bail records are persisted before notification fires.
    # Import here to avoid circular imports.
    try:
        from utils import send_bail_notification
        from run import mail as _mail
    except Exception:
        _mail = None

    for item in created:
        try:
            send_bail_notification(
                district=district,
                accused_name=item['name'],
                fir_label=f"अभियुक्त ID:{item['accused_id']}",
                bail_type='temporary',
                bail_start=None,
                bail_end=None,
                bail_remark='Excel बल्क जमानत स्वीकृति',
                bail_rating=0,
                approved_by_name=f'Batch #{batch_id}',
                approved_by_id=resolved_by_uid,
                mail_instance=_mail,
            )
        except Exception as _notif_err:
            import logging
            logging.getLogger(__name__).error(
                f"[Notify] Excel batch confirm notification error for "
                f"accused {item['accused_id']}: {_notif_err}"
            )

    # ── WhatsApp: ONE consolidated message for the whole batch ────────────────
    # Unlike push/email/in-app above (sent per accused), WhatsApp combines
    # every accused approved in this batch into a single message per
    # recipient, so an admin approving 20 bails in one Excel upload doesn't
    # flood district officers with 20 separate WhatsApp messages.
    if created:
        try:
            from whatsapp_service import send_bail_whatsapp_notification
            wa_bails = [
                {
                    "accused_name": item['name'],
                    "fir_label": f"अभियुक्त ID:{item['accused_id']}",
                    "bail_type": "temporary",
                    "bail_start": None,
                    "bail_end": None,
                    "bail_remark": "Excel बल्क जमानत स्वीकृति",
                    "bail_rating": 0,
                }
                for item in created
            ]
            wa_result = send_bail_whatsapp_notification(
                district=district,
                bails=wa_bails,
                approved_by_name=f'Batch #{batch_id} ({resolved_by_uid})',
                approved_by_id=resolved_by_uid,
            )
            import logging
            logging.getLogger(__name__).info(
                f"[WhatsApp] Batch #{batch_id} consolidated notify: "
                f"✓{wa_result.get('sent', 0)} ✗{wa_result.get('failed', 0)}"
            )
        except Exception as _wa_err:
            import logging
            logging.getLogger(__name__).error(
                f"[WhatsApp] Batch #{batch_id} consolidated notify error: {_wa_err}"
            )

    return {'created': created, 'skipped': skipped}


# ── Pending photo/document completion ────────────────────────────────────────

def list_pending_photo_bails(district: str):
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT abh.id AS bail_id, abh.accused_id, abh.bail_start_date, abh.court_name,
               abh.jail_date, abh.release_date, abh.approved_at,
               a.name, a.fathers_name, a.photo_url,
               GROUP_CONCAT(DISTINCT f.fir_number SEPARATOR ', ') AS fir_numbers,
               GROUP_CONCAT(DISTINCT f.thana SEPARATOR ', ') AS thanas
        FROM accused_bail_history abh
        JOIN accused a ON a.id = abh.accused_id
        LEFT JOIN accused_bail_fir abf ON abf.bail_id = abh.id
        LEFT JOIN fir_cases f ON f.id = abf.fir_id
        WHERE abh.photo_status = 'pending' AND abh.status = 'ACTIVE'
        AND EXISTS (
            SELECT 1 FROM accused_bail_fir x JOIN fir_cases fx ON fx.id = x.fir_id
            WHERE x.bail_id = abh.id AND fx.district = %s
        )
        GROUP BY abh.id
        ORDER BY abh.approved_at DESC
    """, (district,))
    rows = cursor.fetchall()
    cursor.close(); conn.close()
    return rows


def complete_bail_photo(bail_id: int, district: str, uid: int,
                         photo_data: str = None, doc_file=None):
    """
    Upload the (optionally later-captured) geo-tagged photo and/or
    supporting document for a bail record that was created via Excel bulk
    import (photo_status='pending'), and mark it complete.
    """
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT abh.*, a.name, a.photo_url AS accused_photo_url
        FROM accused_bail_history abh JOIN accused a ON a.id = abh.accused_id
        WHERE abh.id=%s
        AND EXISTS (
            SELECT 1 FROM accused_bail_fir x JOIN fir_cases fx ON fx.id = x.fir_id
            WHERE x.bail_id = abh.id AND fx.district = %s
        )
    """, (bail_id, district))
    bail = cursor.fetchone()
    if not bail:
        cursor.close(); conn.close()
        return False, 'जमानत रिकॉर्ड नहीं मिला या आपके जिले का नहीं है।'
    if bail['photo_status'] == 'complete':
        cursor.close(); conn.close()
        return False, 'यह रिकॉर्ड पहले ही पूर्ण किया जा चुका है।'

    photo_url, photo_public_id = None, None
    if photo_data and photo_data.startswith('data:image'):
        photo_url, photo_public_id = upload_image(photo_data, folder='accused_bail_photos')

    doc_url, doc_public_id, doc_res_type = None, None, 'raw'
    if doc_file and getattr(doc_file, 'filename', None):
        doc_url, doc_public_id, doc_res_type = upload_document(doc_file, folder='accused_bail_docs')

    if not photo_url and not doc_url:
        cursor.close(); conn.close()
        return False, 'कृपया फ़ोटो या दस्तावेज़ में से कम से कम एक अपलोड करें।'

    set_clauses, params = [], []
    if photo_url:
        set_clauses += ["bail_photo_url=%s", "bail_photo_public_id=%s"]
        params += [photo_url, photo_public_id]
    if doc_url:
        set_clauses += ["bail_document_url=%s", "bail_document_public_id=%s", "bail_document_resource_type=%s"]
        params += [doc_url, doc_public_id, doc_res_type]
    set_clauses.append("photo_status='complete'")
    params.append(bail_id)
    cursor.execute(f"UPDATE accused_bail_history SET {', '.join(set_clauses)} WHERE id=%s", params)

    # Mirror onto accused's current-bail columns + profile photo if missing
    acc_clauses, acc_params = [], []
    if photo_url:
        acc_clauses += ["bail_photo_url=%s", "bail_photo_public_id=%s"]
        acc_params += [photo_url, photo_public_id]
        if not bail['accused_photo_url']:
            acc_clauses += ["photo_url=%s", "photo_public_id=%s", "profile_status='complete'"]
            acc_params += [photo_url, photo_public_id]
            cursor.execute("UPDATE accused_photos SET is_current=0 WHERE accused_id=%s", (bail['accused_id'],))
            cursor.execute("""
                INSERT INTO accused_photos (accused_id, photo_url, photo_public_id, is_current, uploaded_by)
                VALUES (%s,%s,%s,1,%s)
            """, (bail['accused_id'], photo_url, photo_public_id, uid))
    if doc_url:
        acc_clauses += ["bail_documents_url=%s", "bail_documents_public_id=%s"]
        acc_params += [doc_url, doc_public_id]
    if acc_clauses:
        acc_clauses.append("updated_by=%s")
        acc_params.append(uid)
        acc_params.append(bail['accused_id'])
        cursor.execute(f"UPDATE accused SET {', '.join(acc_clauses)} WHERE id=%s", acc_params)

    conn.commit()
    cursor.close()
    conn.close()
    return True, 'फ़ोटो/दस्तावेज़ सफलतापूर्वक जोड़ा गया — जमानत रिकॉर्ड पूर्ण।'

# ══════════════════════════════════════════════════════════════════════════
# Flask route handlers — thin views over the functions above, following the
# same (role='admin'|'super') convention as accused_common.py so admin.py /
# super_admin.py only need one-line route wrappers around these.
# ══════════════════════════════════════════════════════════════════════════

def _bp_for(role):
    return 'admin' if role == 'admin' else 'super'


ALLOWED_EXCEL_EXTENSIONS = ('.xlsx', '.xls')
MAX_EXCEL_FILE_BYTES = 15 * 1024 * 1024  # 15 MB — generous for a few-hundred-row sheet


def handle_bail_excel_upload(role='admin'):
    """
    GET  -> upload form
    POST -> validate + stage_bail_excel() -> redirect to review page
    """
    bp = _bp_for(role)
    district = session.get('district')

    if request.method == 'POST':
        file = request.files.get('bail_excel')
        if not file or not file.filename:
            flash('कृपया एक Excel फ़ाइल चुनें।', 'danger')
            return redirect(url_for(f'{bp}.bail_excel_upload'))

        filename = file.filename
        if not filename.lower().endswith(ALLOWED_EXCEL_EXTENSIONS):
            flash('केवल .xlsx / .xls फ़ाइलें स्वीकार्य हैं।', 'danger')
            return redirect(url_for(f'{bp}.bail_excel_upload'))

        file.seek(0, 2)
        size = file.tell()
        file.seek(0)
        if size > MAX_EXCEL_FILE_BYTES:
            flash('फ़ाइल आकार 15MB से अधिक है — कृपया छोटी फ़ाइल अपलोड करें।', 'danger')
            return redirect(url_for(f'{bp}.bail_excel_upload'))
        if size == 0:
            flash('फ़ाइल खाली है।', 'danger')
            return redirect(url_for(f'{bp}.bail_excel_upload'))

        try:
            batch_id = stage_bail_excel(file, filename, district, session['user_id'])
        except ValueError as e:
            flash(str(e), 'danger')
            return redirect(url_for(f'{bp}.bail_excel_upload'))

        log_activity(session['user_id'], session.get('role'),
                     f"Uploaded bail Excel batch #{batch_id} ('{filename}')",
                     ip=request.remote_addr)
        flash('फ़ाइल सफलतापूर्वक अपलोड व प्रोसेस हो गई — नीचे परिणाम की समीक्षा करें।', 'success')
        return redirect(url_for(f'{bp}.bail_excel_review', batch_id=batch_id))

    batches = list_batches(district)
    return render_template(f'{bp}/bail_excel_upload.html', batches=batches)


def handle_batch_review(batch_id: int, role='admin'):
    """GET -> show every row of a staged batch, grouped by match status."""
    bp = _bp_for(role)
    district = session.get('district')
    batch, rows = get_batch_review(batch_id, district)
    if not batch:
        flash('बैच नहीं मिला।', 'danger')
        return redirect(url_for(f'{bp}.bail_excel_upload'))

    grouped = {'matched': [], 'ambiguous': [], 'not_found': [], 'already_bailed': [], 'fir_not_found': []}
    for r in rows:
        grouped.setdefault(r['match_status'], []).append(r)

    return render_template(f'{bp}/bail_batch_review.html', batch=batch, rows=rows, grouped=grouped)


def handle_resolve_ambiguous(batch_id: int, row_id: int, role='admin'):
    """POST -> admin manually resolved an ambiguous row to a specific accused_id."""
    bp = _bp_for(role)
    district = session.get('district')
    accused_id = request.form.get('accused_id', type=int)
    if not accused_id:
        flash('कृपया सूची में से एक अभियुक्त चुनें।', 'danger')
        return redirect(url_for(f'{bp}.bail_excel_review', batch_id=batch_id))

    ok, msg = resolve_ambiguous_row(batch_id, row_id, accused_id, district)
    flash(msg, 'success' if ok else 'danger')
    return redirect(url_for(f'{bp}.bail_excel_review', batch_id=batch_id))


def handle_batch_confirm(batch_id: int, role='admin'):
    """POST -> create real accused_bail_history rows for every checked row."""
    bp = _bp_for(role)
    district = session.get('district')
    confirmed_row_ids = request.form.getlist('row_ids', type=int)

    if not confirmed_row_ids:
        flash('स्वीकृत करने हेतु कम से कम एक पंक्ति चुनें।', 'warning')
        return redirect(url_for(f'{bp}.bail_excel_review', batch_id=batch_id))

    result = confirm_batch(batch_id, district, confirmed_row_ids, session['user_id'])
    if result.get('error'):
        flash(result['error'], 'danger')
        return redirect(url_for(f'{bp}.bail_excel_upload'))

    log_activity(session['user_id'], session.get('role'),
                 f"Confirmed bail Excel batch #{batch_id}: "
                 f"{len(result['created'])} created, {len(result['skipped'])} skipped",
                 ip=request.remote_addr)

    if result['created']:
        names = ', '.join(c['name'] for c in result['created'][:5])
        more = f" +{len(result['created'])-5} और" if len(result['created']) > 5 else ''
        flash(f"{len(result['created'])} अभियुक्तों की जमानत स्वीकृत की गई: {names}{more}। "
              f"फ़ोटो/दस्तावेज़ 'लंबित फ़ोटो' सूची से पूर्ण करें।", 'success')
    if result['skipped']:
        reasons = '; '.join(f"पंक्ति {s['row']}: {s['reason']}" for s in result['skipped'][:5])
        flash(f"{len(result['skipped'])} पंक्तियाँ छोड़ी गईं — {reasons}", 'warning')

    return redirect(url_for(f'{bp}.bail_pending_photos'))


def handle_batch_discard(batch_id: int, role='admin'):
    """POST -> abandon a staged batch, no bail records created."""
    bp = _bp_for(role)
    district = session.get('district')
    ok, msg = discard_batch(batch_id, district, session['user_id'])
    flash(msg, 'success' if ok else 'danger')
    return redirect(url_for(f'{bp}.bail_excel_upload'))


def handle_pending_photos(role='admin'):
    """GET -> list every bail record created by Excel-bulk import that is
    still waiting for its geo-tagged photo / supporting document."""
    bp = _bp_for(role)
    district = session.get('district')
    pending = list_pending_photo_bails(district)
    return render_template(f'{bp}/bail_pending_photos.html', pending=pending)


def handle_complete_photo(bail_id: int, role='admin'):
    """POST -> attach the captured photo and/or document to a pending record."""
    bp = _bp_for(role)
    district = session.get('district')
    photo_data = request.form.get('bail_photo_data', '').strip() or None
    doc_file = request.files.get('bail_document')

    ok, msg = complete_bail_photo(bail_id, district, session['user_id'],
                                   photo_data=photo_data, doc_file=doc_file)
    flash(msg, 'success' if ok else 'danger')
    if ok:
        log_activity(session['user_id'], session.get('role'),
                     f"Completed pending photo/document for bail record #{bail_id}",
                     ip=request.remote_addr)
    return redirect(url_for(f'{bp}.bail_pending_photos'))