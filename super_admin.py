"""
super_admin.py
==============
Super Admin routes. This system is fully Accused-based (अभियुक्त-आधारित)
— there is no criminal-management module.
"""
from flask import Blueprint, render_template, request, redirect, url_for, session, flash, jsonify
from functools import wraps
from db import get_connection
from utils import (log_activity, upload_image, upload_document, upload_id_card_file,
                   get_notifications, mark_notifications_read, paginate_query,
                   get_accused_bail_alerts, auto_complete_expired_accused_bails)
from accused_common import (get_fir_list, get_fir_detail, get_accused_list,
                             get_accused_detail, upload_accused_excel,
                             create_fir_manual, download_accused_sample_file,
                             approve_accused_bail, revoke_accused_bail,
                             get_bailed_accused_list)
from bail_bulk import (handle_bail_excel_upload, handle_batch_review,
                        handle_resolve_ambiguous, handle_batch_confirm,
                        handle_batch_discard, handle_pending_photos,
                        handle_complete_photo)
from thana_service import (list_thanas, add_thana, update_thana, toggle_thana,
                            delete_thana, bulk_upload_thanas_from_excel)
import logging, csv, io
from datetime import datetime
from werkzeug.security import generate_password_hash
from flask import make_response, send_file

super_bp = Blueprint('super', __name__)
logger   = logging.getLogger(__name__)


def super_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session or session.get('role') not in ('super_admin',):
            flash('पहुँच अस्वीकृत।', 'danger')
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return decorated


# ══════════════════════════════════════════════════════════════════════════════
# DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════

@super_bp.route('/dashboard')
@super_required
def dashboard():
    district = session.get('district')
    auto_complete_expired_accused_bails(district)
    conn   = get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT COUNT(*) as c FROM users WHERE created_by=%s AND role='admin' AND is_active=1", (session['user_id'],))
    admin_count = cursor.fetchone()['c']
    cursor.execute("""
        SELECT COUNT(DISTINCT a.id) as c FROM accused a
        JOIN accused_fir af ON af.accused_id=a.id
        JOIN fir_cases f ON f.id=af.fir_id WHERE f.district=%s
    """, (district,))
    accused_count = cursor.fetchone()['c']
    cursor.execute("""
        SELECT COUNT(DISTINCT a.id) as c FROM accused a
        JOIN accused_fir af ON af.accused_id=a.id
        JOIN fir_cases f ON f.id=af.fir_id
        WHERE f.district=%s AND a.bail_status!='none'
    """, (district,))
    bail_count = cursor.fetchone()['c']
    cursor.execute("""
        SELECT COUNT(DISTINCT a.id) as c FROM accused a
        JOIN accused_fir af ON af.accused_id=a.id
        JOIN fir_cases f ON f.id=af.fir_id
        WHERE f.district=%s AND a.profile_status='pending'
    """, (district,))
    pending_count = cursor.fetchone()['c']
    cursor.execute("SELECT COUNT(*) as c FROM fir_cases WHERE district=%s", (district,))
    fir_count = cursor.fetchone()['c']
    cursor.close(); conn.close()
    alerts = get_accused_bail_alerts(district)
    stats  = {
        'admins': admin_count, 'accused': accused_count,
        'on_bail': bail_count, 'pending': pending_count,
        'firs': fir_count,
    }
    return render_template('super/dashboard.html', stats=stats, alerts=alerts)


@super_bp.route('/admins')
@super_required
def admins():
    district = session.get('district')

    search = request.args.get("search", "").strip()
    page = int(request.args.get("page", 1))
    per_page = int(request.args.get("per_page", 20))

    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    base_query = """
        SELECT u.*, c.name AS created_by_name
        FROM users u
        LEFT JOIN users c ON c.id = u.created_by
        WHERE u.created_by=%s
        AND u.role='admin'
    """

    params = [session["user_id"]]

    if search:
        base_query += """
        AND (
            u.name LIKE %s
            OR u.user_id LIKE %s
            OR u.designation LIKE %s
            OR u.contact LIKE %s
        )
        """
        like = f"%{search}%"
        params.extend([like, like, like, like])

    base_query += " ORDER BY u.created_at DESC"

    admins_list, total, total_pages = paginate_query(
        cursor,
        base_query,
        params,
        page,
        per_page
    )

    cursor.close()
    conn.close()

    return render_template(
        "super/admins.html",
        admins=admins_list,
        search=search,
        page=page,
        per_page=per_page,
        total=total,
        total_pages=total_pages,
    )

@super_bp.route('/create-admin', methods=['GET', 'POST'])
@super_required
def create_admin():
    if request.method == 'POST':
        uid         = request.form.get('user_id', '').strip().upper()
        name        = request.form.get('name', '').strip()
        designation = request.form.get('designation', '').strip()
        contact     = request.form.get('contact', '').strip()
        email       = request.form.get('email', '').strip()
        district    = session.get('district')
        address     = request.form.get('address', '').strip()
        password    = request.form.get('password', '').strip()

        if not all([uid, name, email, password]):
            flash('User ID, नाम, ईमेल और पासवर्ड आवश्यक हैं।', 'danger')
            return render_template('super/create_admin.html')

        conn   = get_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id FROM users WHERE user_id=%s OR email=%s", (uid, email))
        if cursor.fetchone():
            flash('User ID या ईमेल पहले से मौजूद है।', 'danger')
            cursor.close(); conn.close()
            return render_template('super/create_admin.html')

        cursor.execute("""
            INSERT INTO users (user_id,name,designation,contact,email,district,address,
                               password_hash,role,created_by,is_active)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,'admin',%s,1)
        """, (uid, name, designation, contact, email, district, address,
              generate_password_hash(password), session['user_id']))
        conn.commit()
        log_activity(session['user_id'], session['role'], f"Created admin {uid}", ip=request.remote_addr)
        flash(f'व्यवस्थापक {name} ({uid}) सफलतापूर्वक बनाया गया।', 'success')
        cursor.close(); conn.close()
        return redirect(url_for('super.admins'))
    return render_template('super/create_admin.html',form={})

@super_bp.route('/download-admin-sample')
@super_required
def download_admin_sample():
    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow([
        "user_id",
        "name",
        "designation",
        "contact",
        "email",
        "address",
        "password"
    ])

    writer.writerow([
        "ADM001",
        "Rahul Sharma",
        "Head Constable",
        "9876543210",
        "rahul@example.com",
        "Police Station Mirzapur",
        "Admin@123"
    ])

    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = "attachment; filename=admin_sample.csv"
    response.headers["Content-Type"] = "text/csv"
    return response


@super_bp.route('/admin/<int:uid>/toggle', methods=['POST'])
@super_required
def toggle_admin(uid):
    conn   = get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT id,is_active,user_id FROM users WHERE id=%s AND created_by=%s", (uid, session['user_id']))
    user = cursor.fetchone()
    if not user:
        flash('व्यवस्थापक नहीं मिला।', 'danger')
        cursor.close(); conn.close()
        return redirect(url_for('super.admins'))
    new_status = 0 if user['is_active'] else 1
    cursor.execute("UPDATE users SET is_active=%s WHERE id=%s", (new_status, uid))
    conn.commit()
    status_text = 'सक्रिय' if new_status else 'निष्क्रिय'
    flash(f"व्यवस्थापक {user['user_id']} को {status_text} किया गया।", 'success')
    cursor.close(); conn.close()
    return redirect(url_for('super.admins'))


@super_bp.route('/upload-admins', methods=['GET', 'POST'])
@super_required
def upload_admins():
    if request.method == 'POST':
        file = request.files.get('csv_file')
        if not file or not file.filename.endswith('.csv'):
            flash('कृपया CSV फ़ाइल अपलोड करें।', 'danger')
            return render_template('super/upload_admins.html')
        stream  = io.StringIO(file.stream.read().decode('utf-8'))
        reader  = csv.DictReader(stream)
        conn    = get_connection()
        cursor  = conn.cursor(dictionary=True)
        success, failed = 0, []
        for i, row in enumerate(reader, start=2):
            uid  = (row.get('user_id','') or '').strip().upper()
            name = (row.get('name','') or '').strip()
            email= (row.get('email','') or '').strip()
            pwd  = (row.get('password','') or '').strip()
            if not all([uid, name, email, pwd]):
                failed.append(f"पंक्ति {i}: आवश्यक फ़ील्ड खाली"); continue
            cursor.execute("SELECT id FROM users WHERE user_id=%s OR email=%s", (uid, email))
            if cursor.fetchone():
                failed.append(f"पंक्ति {i} ({uid}): पहले से मौजूद"); continue
            try:
                cursor.execute("""
                    INSERT INTO users (user_id,name,designation,contact,email,district,address,
                                       password_hash,role,created_by,is_active)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,'admin',%s,1)
                """, (uid, name,
                      (row.get('designation','') or '').strip(),
                      (row.get('contact','') or '').strip(),
                      email, session.get('district'),
                      (row.get('address','') or '').strip(),
                      generate_password_hash(pwd), session['user_id']))
                success += 1
            except Exception as e:
                failed.append(f"पंक्ति {i}: {str(e)[:80]}")
        conn.commit(); cursor.close(); conn.close()
        flash(f"{success} व्यवस्थापक जोड़े गए। असफल: {len(failed)}", 'success' if not failed else 'warning')
        for m in failed[:10]: flash(m, 'warning')
        return redirect(url_for('super.admins'))
    return render_template('super/upload_admins.html')


@super_bp.route('/notifications')
@super_required
def notifications():
    notifs = get_notifications(session['user_id'], limit=50)
    mark_notifications_read(session['user_id'])
    return render_template('super/notifications.html', notifications=notifs)


# ══════════════════════════════════════════════════════════════════════════════
# थाना (POLICE STATION) DIRECTORY — WhatsApp/email bail-approval routing
# ══════════════════════════════════════════════════════════════════════════════

@super_bp.route('/thana')
@super_required
def thana_list():
    district = session.get('district')
    search = request.args.get('search', '').strip()
    thanas = list_thanas(district, search)
    return render_template('super/thana_list.html', thanas=thanas, search=search)


@super_bp.route('/thana/add', methods=['POST'])
@super_required
def thana_add():
    district = session.get('district')
    thana_name = request.form.get('thana_name', '').strip()
    contact    = request.form.get('contact', '').strip()
    email      = request.form.get('email', '').strip()
    ok, err = add_thana(district, thana_name, contact, email, session['user_id'])
    if ok:
        log_activity(session['user_id'], session['role'], f"Added/updated थाना {thana_name}", ip=request.remote_addr)
        flash(f'थाना "{thana_name}" सफलतापूर्वक जोड़ा गया।', 'success')
    else:
        flash(err or 'थाना जोड़ने में त्रुटि।', 'danger')
    return redirect(url_for('super.thana_list'))


@super_bp.route('/thana/<int:thana_id>/edit', methods=['POST'])
@super_required
def thana_edit(thana_id):
    district = session.get('district')
    thana_name = request.form.get('thana_name', '').strip()
    contact    = request.form.get('contact', '').strip()
    email      = request.form.get('email', '').strip()
    ok, err = update_thana(thana_id, district, thana_name, contact, email)
    if ok:
        log_activity(session['user_id'], session['role'], f"Updated थाना #{thana_id}", ip=request.remote_addr)
        flash('थाना विवरण अपडेट किया गया।', 'success')
    else:
        flash(err or 'अपडेट करने में त्रुटि।', 'danger')
    return redirect(url_for('super.thana_list'))


@super_bp.route('/thana/<int:thana_id>/toggle', methods=['POST'])
@super_required
def thana_toggle(thana_id):
    district = session.get('district')
    if toggle_thana(thana_id, district):
        flash('थाना की स्थिति बदल दी गई।', 'success')
    else:
        flash('थाना नहीं मिला।', 'danger')
    return redirect(url_for('super.thana_list'))


@super_bp.route('/thana/<int:thana_id>/delete', methods=['POST'])
@super_required
def thana_delete(thana_id):
    district = session.get('district')
    if delete_thana(thana_id, district):
        log_activity(session['user_id'], session['role'], f"Deleted थाना #{thana_id}", ip=request.remote_addr)
        flash('थाना हटा दिया गया।', 'success')
    else:
        flash('थाना नहीं मिला।', 'danger')
    return redirect(url_for('super.thana_list'))


@super_bp.route('/thana/upload', methods=['GET', 'POST'])
@super_required
def thana_upload():
    if request.method == 'POST':
        file = request.files.get('excel_file')
        if not file or not file.filename:
            flash('कृपया Excel फ़ाइल चुनें।', 'danger')
            return render_template('super/thana_upload.html')
        fname = file.filename.lower()
        if not (fname.endswith('.xlsx') or fname.endswith('.xls')):
            flash('केवल .xlsx या .xls फ़ाइल अपलोड करें।', 'danger')
            return render_template('super/thana_upload.html')

        result = bulk_upload_thanas_from_excel(session.get('district'), file, session['user_id'])
        flash(f"{result['success']} थाना जोड़े/अपडेट किए गए। असफल: {len(result['failed'])}",
              'success' if not result['failed'] else 'warning')
        for m in result['failed'][:10]:
            flash(m, 'warning')
        log_activity(session['user_id'], session['role'],
                     f"Bulk uploaded {result['success']} थाना", ip=request.remote_addr)
        return redirect(url_for('super.thana_list'))
    return render_template('super/thana_upload.html')


@super_bp.route('/thana/download-sample')
@super_required
def thana_download_sample():
    """Return a sample .xlsx matching the format bulk_upload_thanas_from_excel expects."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'थाना सूची'

        headers = ['thana_name', 'contact', 'email']
        ws.append(headers)
        header_fill = PatternFill("solid", fgColor="1a73e8")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        ws.append(['थाना कोतवाली', '9876543210', 'kotwali@example.com'])
        ws.append(['थाना देहात', '9876500000', 'dehat.thana@example.com'])

        for i, w in enumerate([25, 18, 30], 1):
            ws.column_dimensions[chr(64 + i)].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True, download_name='thana_upload_sample.xlsx'
        )
    except ImportError:
        flash('openpyxl install करें: pip install openpyxl', 'danger')
        return redirect(url_for('super.thana_list'))


# ══════════════════════════════════════════════════════════════════════════════
# ACCUSED & FIR ROUTES
# ══════════════════════════════════════════════════════════════════════════════

@super_bp.route('/accused')
@super_required
def accused_list():
    return get_accused_list(role='super')


@super_bp.route('/accused/<int:accused_id>')
@super_required
def accused_detail(accused_id):
    return get_accused_detail(accused_id)


@super_bp.route('/accused/<int:accused_id>/upload-photo', methods=['POST'])
@super_required
def upload_accused_photo(accused_id):
    photo = request.files.get('photo')
    if not photo or not photo.filename:
        flash('फ़ोटो चुनें।', 'danger')
        return redirect(url_for('super.accused_detail', accused_id=accused_id))
    url, pub_id = upload_image(photo, folder='accused_photos')
    if url:
        conn = get_connection(); cursor = conn.cursor()
        cursor.execute("UPDATE accused_photos SET is_current=0 WHERE accused_id=%s", (accused_id,))
        cursor.execute("""
            INSERT INTO accused_photos (accused_id,photo_url,photo_public_id,is_current,uploaded_by)
            VALUES (%s,%s,%s,1,%s)
        """, (accused_id, url, pub_id, session['user_id']))
        cursor.execute("UPDATE accused SET photo_url=%s,photo_public_id=%s,profile_status='complete' WHERE id=%s",
                       (url, pub_id, accused_id))
        conn.commit(); cursor.close(); conn.close()
        flash('फ़ोटो अपलोड सफल।', 'success')
    else:
        flash('फ़ोटो अपलोड असफल।', 'danger')
    return redirect(url_for('super.accused_detail', accused_id=accused_id))


@super_bp.route('/accused/<int:accused_id>/approve-bail', methods=['GET', 'POST'])
@super_required
def approve_bail(accused_id):
    """
    जमानत स्वीकृत करना अब केवल जिला Admin का कार्य है — Super Admin की भूमिका
    निगरानी (view), Admin/FIR/Accused निर्माण, और Excel बल्क अपलोड तक सीमित है।
    """
    flash('जमानत स्वीकृत करना केवल जिला Admin द्वारा किया जा सकता है। Super Admin केवल देख सकते हैं।', 'danger')
    return redirect(url_for('super.accused_detail', accused_id=accused_id))


@super_bp.route('/accused/<int:accused_id>/revoke-bail', methods=['POST'])
@super_required
def revoke_bail_accused(accused_id):
    flash('जमानत रद्द करना केवल जिला Admin द्वारा किया जा सकता है।', 'danger')
    return redirect(url_for('super.accused_detail', accused_id=accused_id))


@super_bp.route('/bailed-accused')
@super_required
def bailed_accused():
    return get_bailed_accused_list(role='super')


# ══════════════════════════════════════════════════════════════════════════════
# BAIL EXCEL BULK APPROVAL — same feature-set as admin, super_admin scope
# ══════════════════════════════════════════════════════════════════════════════

@super_bp.route('/bail-excel/upload', methods=['GET', 'POST'])
@super_required
def bail_excel_upload():
    return handle_bail_excel_upload(role='super')


@super_bp.route('/bail-excel/batch/<int:batch_id>')
@super_required
def bail_excel_review(batch_id):
    return handle_batch_review(batch_id, role='super')


@super_bp.route('/bail-excel/batch/<int:batch_id>/row/<int:row_id>/resolve', methods=['POST'])
@super_required
def bail_excel_resolve(batch_id, row_id):
    return handle_resolve_ambiguous(batch_id, row_id, role='super')


@super_bp.route('/bail-excel/batch/<int:batch_id>/confirm', methods=['POST'])
@super_required
def bail_excel_confirm(batch_id):
    return handle_batch_confirm(batch_id, role='super')


@super_bp.route('/bail-excel/batch/<int:batch_id>/discard', methods=['POST'])
@super_required
def bail_excel_discard(batch_id):
    return handle_batch_discard(batch_id, role='super')


@super_bp.route('/bail-pending-photos')
@super_required
def bail_pending_photos():
    return handle_pending_photos(role='super')


@super_bp.route('/bail-pending-photos/<int:bail_id>/complete', methods=['POST'])
@super_required
def bail_complete_photo(bail_id):
    flash('फ़ोटो/दस्तावेज़ पूर्ण करना केवल जिला Admin द्वारा किया जा सकता है। Super Admin केवल देख सकते हैं।', 'danger')
    return redirect(url_for('super.bail_pending_photos'))


@super_bp.route('/fir')
@super_required
def fir_list():
    return get_fir_list(role='super')


@super_bp.route('/fir/<int:fir_id>')
@super_required
def fir_detail(fir_id):
    return get_fir_detail(fir_id)


@super_bp.route('/fir/add', methods=['GET', 'POST'])
@super_required
def add_fir():
    return create_fir_manual(role='super')


@super_bp.route('/upload-accused', methods=['GET', 'POST'])
@super_required
def upload_accused():
    return upload_accused_excel(role='super')


@super_bp.route('/download-accused-sample')
@super_required
def download_accused_sample():
    return download_accused_sample_file()