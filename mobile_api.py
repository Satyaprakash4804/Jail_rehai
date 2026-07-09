# -*- coding: utf-8 -*-
"""
mobile_api.py
=============
JSON API layer for the Jail Rehai FLUTTER app.

This file is 100% ADDITIVE — it does not modify, remove, or import-break
any existing route in run.py / admin.py / super_admin.py / master.py /
auth.py. It is registered as one extra blueprint in run.py:

    from mobile_api import mobile_bp
    app.register_blueprint(mobile_bp, url_prefix='/api/mobile')

AUTH MODEL
----------
Same as the web app: Flask server-side session cookie. The Flutter app
uses `dio` + a persistent cookie jar, so after /api/mobile/auth/login sets
the session cookie, every subsequent request automatically carries it —
exactly like a browser. This means:
  * zero changes to auth.py / login security model
  * FCM save-token / delete-token endpoints already in run.py work as-is
  * activity_logs / log_activity keeps working unmodified

Every endpoint returns JSON: {"success": bool, ...}. Errors use standard
HTTP status codes (401 unauthenticated, 403 wrong role, 404 not found,
400 validation, 500 server error).
"""
from flask import Blueprint, request, session, jsonify, send_file
from functools import wraps
from datetime import datetime, date
import logging
import io

from db import get_connection
from utils import (log_activity, upload_image, upload_document,
                    get_notifications, mark_notifications_read,
                    paginate_query, get_accused_bail_alerts,
                    auto_complete_expired_accused_bails, generate_otp,
                    send_bail_notification)
from accused_common import (find_or_create_accused, upsert_accused_fir,
                             parse_accused_list, normalize_name,
                             download_accused_sample_file)
from bail_bulk import (stage_bail_excel, get_batch_review,
                        resolve_ambiguous_row, discard_batch, list_batches,
                        confirm_batch, list_pending_photo_bails,
                        complete_bail_photo, MATCH_STATUS_LABELS)
from werkzeug.security import check_password_hash, generate_password_hash

mobile_bp = Blueprint('mobile_api', __name__)
logger = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════════════════════
# Helpers
# ══════════════════════════════════════════════════════════════════════════

def ok(data=None, **extra):
    body = {"success": True}
    if data is not None:
        body["data"] = data
    body.update(extra)
    return jsonify(body)


def err(message, status=400, **extra):
    body = {"success": False, "error": message}
    body.update(extra)
    return jsonify(body), status


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return err('Authentication required.', 401)
        return f(*args, **kwargs)
    return decorated


def role_required(*roles):
    def wrapper(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if 'user_id' not in session:
                return err('Authentication required.', 401)
            if session.get('role') not in roles:
                return err('Access denied for this role.', 403)
            return f(*args, **kwargs)
        return decorated
    return wrapper


def current_user_payload():
    return {
        "id": session.get("user_id"),
        "user_id": session.get("user_uid"),
        "name": session.get("name"),
        "role": session.get("role"),
        "email": session.get("email"),
        "district": session.get("district"),
        "designation": session.get("designation"),
    }


def bail_bp():
    """district-scoping bp string used by bail_bulk helper funcs (admin/super share same scope key)"""
    return session.get('role')


def jsonify_dates(row):
    """MySQL DATE/DATETIME -> ISO strings so json.dumps never chokes."""
    for k, v in list(row.items()):
        if isinstance(v, (datetime, date)):
            row[k] = v.isoformat()
    return row


def rows_iso(rows):
    return [jsonify_dates(dict(r)) for r in rows]


# ══════════════════════════════════════════════════════════════════════════
# AUTH
# ══════════════════════════════════════════════════════════════════════════

@mobile_bp.route('/auth/login', methods=['POST'])
def login():
    data = request.get_json(silent=True) or {}
    user_id = (data.get('user_id') or '').strip()
    password = (data.get('password') or '').strip()
    if not user_id or not password:
        return err('User ID and password are required.', 400)

    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM users WHERE user_id=%s", (user_id,))
    user = cursor.fetchone()
    cursor.close()
    conn.close()

    if not user:
        return err('Invalid User ID or password.', 401)
    if not user['is_active']:
        return err('Your account has been revoked. Contact administrator.', 403)
    if not check_password_hash(user['password_hash'], password):
        return err('Invalid User ID or password.', 401)

    session['user_id'] = user['id']
    session['user_uid'] = user['user_id']
    session['name'] = user['name']
    session['role'] = user['role']
    session['email'] = user['email']
    session['district'] = user['district']
    session['designation'] = user['designation']
    session.permanent = True

    log_activity(user['id'], user['role'], 'User logged in (mobile)', ip=request.remote_addr)
    return ok(current_user_payload())


@mobile_bp.route('/auth/logout', methods=['POST'])
@login_required
def logout():
    user_id = session.get('user_id')
    role = session.get('role')
    log_activity(user_id, role, 'User logged out (mobile)', ip=request.remote_addr)
    try:
        from fcm_service import delete_user_tokens
        delete_user_tokens(user_id)
    except Exception as e:
        logger.warning(f"[FCM] Could not delete tokens on logout: {e}")
    session.clear()
    return ok()


@mobile_bp.route('/auth/me')
@login_required
def me():
    return ok(current_user_payload())


@mobile_bp.route('/auth/forgot-password', methods=['POST'])
def forgot_password():
    data = request.get_json(silent=True) or {}
    email = (data.get('email') or '').strip()
    if not email:
        return err('Email is required.', 400)

    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM users WHERE email=%s", (email,))
    user = cursor.fetchone()
    if not user:
        cursor.close(); conn.close()
        return err('No account found with this email.', 404)

    otp = generate_otp()
    from datetime import timedelta
    expiry = datetime.now() + timedelta(minutes=10)
    cursor.execute("UPDATE users SET otp_code=%s, otp_expiry=%s WHERE id=%s",
                   (otp, expiry, user['id']))
    conn.commit()
    cursor.close(); conn.close()

    try:
        from flask_mail import Message
        from flask import current_app
        mail = current_app.extensions['mail']
        msg = Message('Password Reset OTP - Jail Rehai', recipients=[email],
                      sender='noreply@jailrehai.gov.in')
        msg.body = (f"Dear {user['name']},\n\nYour OTP for password reset is: {otp}\n\n"
                    f"Valid for 10 minutes. Do not share this OTP.\n\nRegards,\nJail Rehai System")
        mail.send(msg)
        return ok(message="OTP sent to your email.")
    except Exception as e:
        logger.error(f"Mail error: {e}")
        # dev fallback so testing works even without SMTP configured
        return ok(message="OTP generated (mail not configured).", dev_otp=otp)


@mobile_bp.route('/auth/verify-otp', methods=['POST'])
def verify_otp():
    data = request.get_json(silent=True) or {}
    email = (data.get('email') or '').strip()
    otp = (data.get('otp') or '').strip()
    new_password = (data.get('new_password') or '').strip()
    confirm_password = (data.get('confirm_password') or '').strip()

    if new_password != confirm_password:
        return err('Passwords do not match.', 400)

    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM users WHERE email=%s", (email,))
    user = cursor.fetchone()

    if not user or user['otp_code'] != otp or datetime.now() > user['otp_expiry']:
        cursor.close(); conn.close()
        return err('Invalid or expired OTP.', 400)

    cursor.execute(
        "UPDATE users SET password_hash=%s, otp_code=NULL, otp_expiry=NULL WHERE id=%s",
        (generate_password_hash(new_password), user['id']))
    conn.commit()
    cursor.close(); conn.close()
    return ok(message='Password reset successfully. Please login.')


@mobile_bp.route('/auth/change-password', methods=['POST'])
@login_required
def change_password():
    data = request.get_json(silent=True) or {}
    current = (data.get('current_password') or '').strip()
    new_pass = (data.get('new_password') or '').strip()
    confirm = (data.get('confirm_password') or '').strip()

    if new_pass != confirm:
        return err('Passwords do not match.', 400)

    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM users WHERE id=%s", (session['user_id'],))
    user = cursor.fetchone()
    if not check_password_hash(user['password_hash'], current):
        cursor.close(); conn.close()
        return err('Current password is incorrect.', 400)
    cursor.execute("UPDATE users SET password_hash=%s WHERE id=%s",
                   (generate_password_hash(new_pass), session['user_id']))
    conn.commit()
    cursor.close(); conn.close()
    return ok(message='Password updated successfully.')


# ══════════════════════════════════════════════════════════════════════════
# DASHBOARD (role-aware, one endpoint for all 3 roles)
# ══════════════════════════════════════════════════════════════════════════

@mobile_bp.route('/dashboard')
@login_required
def dashboard():
    role = session.get('role')
    district = session.get('district')
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    if role == 'master':
        cursor.execute("SELECT COUNT(*) as c FROM users WHERE role='super_admin'")
        super_count = cursor.fetchone()['c']
        cursor.execute("SELECT COUNT(*) as c FROM users WHERE role='admin'")
        admin_count = cursor.fetchone()['c']
        cursor.execute("SELECT COUNT(*) as c FROM users WHERE is_active=0")
        revoked_count = cursor.fetchone()['c']
        cursor.execute("SELECT COUNT(*) as c FROM activity_logs")
        log_count = cursor.fetchone()['c']
        cursor.close(); conn.close()
        return ok({'stats': {'super_admins': super_count, 'admins': admin_count,
                              'revoked': revoked_count, 'logs': log_count}, 'alerts': []})

    auto_complete_expired_accused_bails(district)

    cursor.execute("""
        SELECT COUNT(DISTINCT a.id) as c FROM accused a
        JOIN accused_fir af ON af.accused_id=a.id
        JOIN fir_cases f ON f.id=af.fir_id WHERE f.district=%s
    """, (district,))
    accused_count = cursor.fetchone()['c']
    cursor.execute("""
        SELECT COUNT(DISTINCT a.id) as c FROM accused a
        JOIN accused_fir af ON af.accused_id=a.id
        JOIN fir_cases f ON f.id=af.fir_id
        WHERE f.district=%s AND a.profile_status='pending'
    """, (district,))
    pending_count = cursor.fetchone()['c']
    cursor.execute("""
        SELECT COUNT(DISTINCT a.id) as c FROM accused a
        JOIN accused_fir af ON af.accused_id=a.id
        JOIN fir_cases f ON f.id=af.fir_id
        WHERE f.district=%s AND a.profile_status='complete'
    """, (district,))
    complete_count = cursor.fetchone()['c']
    cursor.execute("""
        SELECT COUNT(DISTINCT a.id) as c FROM accused a
        JOIN accused_fir af ON af.accused_id=a.id
        JOIN fir_cases f ON f.id=af.fir_id
        WHERE f.district=%s AND a.bail_status!='none'
    """, (district,))
    bail_count = cursor.fetchone()['c']
    cursor.execute("SELECT COUNT(*) as c FROM fir_cases WHERE district=%s", (district,))
    fir_count = cursor.fetchone()['c']

    stats = {'accused': accused_count, 'pending': pending_count,
             'complete': complete_count, 'on_bail': bail_count, 'firs': fir_count}

    if role == 'super_admin':
        cursor.execute("SELECT COUNT(*) as c FROM users WHERE created_by=%s AND role='admin' AND is_active=1",
                       (session['user_id'],))
        stats['admins'] = cursor.fetchone()['c']
        cursor.execute("SELECT COUNT(*) as c FROM thana_master WHERE district=%s AND is_active=1", (district,))
        stats['thanas'] = cursor.fetchone()['c']

    cursor.close(); conn.close()
    alerts = rows_iso(get_accused_bail_alerts(district))
    return ok({'stats': stats, 'alerts': alerts})


# ══════════════════════════════════════════════════════════════════════════
# NOTIFICATIONS
# ══════════════════════════════════════════════════════════════════════════

@mobile_bp.route('/notifications')
@login_required
def notifications():
    notifs = rows_iso(get_notifications(session['user_id'], limit=100))
    return ok({'notifications': notifs})


@mobile_bp.route('/notifications/mark-read', methods=['POST'])
@login_required
def notifications_mark_read():
    mark_notifications_read(session['user_id'])
    return ok()


@mobile_bp.route('/notifications/count')
@login_required
def notifications_count():
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT COUNT(*) AS c FROM notifications WHERE user_id=%s AND is_read=0",
                   (session['user_id'],))
    c = cursor.fetchone()['c']
    cursor.close(); conn.close()
    return ok({'count': c})


# ══════════════════════════════════════════════════════════════════════════
# ACCUSED
# ══════════════════════════════════════════════════════════════════════════

@mobile_bp.route('/accused')
@role_required('admin', 'super_admin')
def accused_list():
    district = session.get('district')
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 25))
    search = request.args.get('search', '').strip()
    thana_f = request.args.get('thana', '').strip()
    fir_f = request.args.get('fir', '').strip()
    status_f = request.args.get('status', '').strip()

    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    conditions = ["f.district=%s"]
    params = [district]
    if thana_f:
        conditions.append("f.thana LIKE %s"); params.append(f'%{thana_f}%')
    if fir_f:
        conditions.append("f.fir_number LIKE %s"); params.append(f'%{fir_f}%')
    if status_f in ('pending', 'complete'):
        conditions.append("a.profile_status=%s"); params.append(status_f)
    if search:
        conditions.append("(a.name LIKE %s OR a.fathers_name LIKE %s)")
        like = f'%{search}%'; params += [like, like]
    where = " AND ".join(conditions)

    base_q = f"""
        SELECT DISTINCT
            a.id, a.name, a.fathers_name, a.photo_url, a.profile_status, a.bail_status,
            (SELECT COUNT(*) FROM accused_fir af2 WHERE af2.accused_id=a.id) AS fir_count,
            (SELECT GROUP_CONCAT(DISTINCT f2.thana ORDER BY f2.thana SEPARATOR ', ')
             FROM accused_fir af3 JOIN fir_cases f2 ON f2.id=af3.fir_id
             WHERE af3.accused_id=a.id AND f2.district=%s) AS thanas,
            MAX(af.in_arrested) AS ever_arrested
        FROM accused a
        JOIN accused_fir af ON af.accused_id = a.id
        JOIN fir_cases f ON f.id = af.fir_id
        WHERE {where}
        GROUP BY a.id
        ORDER BY a.name
    """
    params_full = [district] + params
    rows, total, total_pages = paginate_query(cursor, base_q, params_full, page, per_page)

    cursor.execute("SELECT DISTINCT thana FROM fir_cases WHERE district=%s ORDER BY thana", (district,))
    thanas = [r['thana'] for r in cursor.fetchall()]
    cursor.close(); conn.close()

    return ok({'accused': rows_iso(rows), 'page': page, 'total': total,
               'total_pages': total_pages, 'per_page': per_page, 'thanas': thanas})


@mobile_bp.route('/accused/<int:accused_id>')
@role_required('admin', 'super_admin')
def accused_detail(accused_id):
    district = session.get('district')
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM accused WHERE id=%s", (accused_id,))
    accused = cursor.fetchone()
    if not accused:
        cursor.close(); conn.close()
        return err('Accused not found.', 404)

    cursor.execute("""
        SELECT f.*, af.in_total_accused, af.in_fir_accused, af.in_arrested, af.in_cs_accused
        FROM accused_fir af JOIN fir_cases f ON f.id = af.fir_id
        WHERE af.accused_id = %s AND f.district = %s ORDER BY f.fir_number
    """, (accused_id, district))
    firs = cursor.fetchall()

    cursor.execute("SELECT * FROM accused_photos WHERE accused_id=%s ORDER BY uploaded_at DESC", (accused_id,))
    photos = cursor.fetchall()

    is_arrested = any(f.get('in_arrested') for f in firs)
    arrest_firs = [f for f in firs if f.get('in_arrested')]
    has_active_bail = bool(accused.get('bail_status') and accused['bail_status'] != 'none')

    cursor.execute("""
        SELECT abh.*, abh.status AS bail_history_status, f.fir_number, f.thana, f.district,
               GROUP_CONCAT(DISTINCT CONCAT(f2.fir_number,'/',f2.thana)
                            ORDER BY f2.fir_number SEPARATOR ', ') AS all_firs,
               COUNT(DISTINCT f2.id) AS fir_count,
               u1.name AS approved_by_name, u2.name AS revoked_by_name
        FROM accused_bail_history abh
        JOIN fir_cases f ON f.id = abh.fir_id
        LEFT JOIN accused_bail_fir abf ON abf.bail_id = abh.id
        LEFT JOIN fir_cases f2 ON f2.id = abf.fir_id
        LEFT JOIN users u1 ON u1.id = abh.approved_by
        LEFT JOIN users u2 ON u2.id = abh.revoked_by
        WHERE abh.accused_id = %s GROUP BY abh.id ORDER BY abh.approved_at DESC
    """, (accused_id,))
    bail_history = cursor.fetchall()
    for b in bail_history:
        if not b.get('all_firs'):
            b['all_firs'] = f"{b['fir_number']}/{b['thana']}"
            b['fir_count'] = 1
    cursor.close(); conn.close()

    return ok({
        'accused': jsonify_dates(dict(accused)), 'firs': rows_iso(firs), 'photos': rows_iso(photos),
        'is_arrested': is_arrested, 'arrest_firs': rows_iso(arrest_firs),
        'has_active_bail': has_active_bail, 'bail_history': rows_iso(bail_history),
    })


@mobile_bp.route('/accused/<int:accused_id>/upload-photo', methods=['POST'])
@role_required('admin', 'super_admin')
def accused_upload_photo(accused_id):
    photo = request.files.get('photo')
    if not photo or not photo.filename:
        return err('Please select a photo.', 400)
    url, pub_id = upload_image(photo, folder='accused_photos')
    if not url:
        return err('Photo upload failed.', 500)
    conn = get_connection(); cursor = conn.cursor()
    cursor.execute("UPDATE accused_photos SET is_current=0 WHERE accused_id=%s", (accused_id,))
    cursor.execute("""INSERT INTO accused_photos (accused_id,photo_url,photo_public_id,is_current,uploaded_by)
                       VALUES (%s,%s,%s,1,%s)""", (accused_id, url, pub_id, session['user_id']))
    cursor.execute("UPDATE accused SET photo_url=%s,photo_public_id=%s,profile_status='complete' WHERE id=%s",
                   (url, pub_id, accused_id))
    conn.commit(); cursor.close(); conn.close()
    log_activity(session['user_id'], session['role'], f"Uploaded photo for accused ID:{accused_id}",
                 ip=request.remote_addr)
    return ok({'photo_url': url})


@mobile_bp.route('/accused/<int:accused_id>/approve-bail', methods=['POST'])
@role_required('admin')  # बल्क Excel के अलावा, एकल जमानत स्वीकृति अब केवल जिला Admin कर सकता है
def accused_approve_bail(accused_id):
    """
    multipart/form-data:
      fir_ids (repeated or comma-separated), bail_type, bail_start_date,
      bail_end_date, bail_remark, bail_rating,
      bail_photo (file, required — geo-tagged capture from the phone camera),
      bail_photo_lat, bail_photo_lng, bail_photo_captured_at (ISO8601),
      bail_document (file, optional)
    """
    import base64
    uid = session['user_id']; role = session['role']
    conn = get_connection(); cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT * FROM accused WHERE id=%s", (accused_id,))
    accused = cursor.fetchone()
    if not accused:
        cursor.close(); conn.close(); return err('Accused not found.', 404)

    cursor.execute("""
        SELECT f.id, f.fir_number, f.thana, f.district
        FROM accused_fir af JOIN fir_cases f ON f.id = af.fir_id
        WHERE af.accused_id=%s AND af.in_arrested=1 ORDER BY f.fir_number
    """, (accused_id,))
    arrest_firs = cursor.fetchall()
    if not arrest_firs:
        cursor.close(); conn.close()
        return err('Bail can only be approved for an accused marked arrested in some FIR.', 400)
    if accused.get('bail_status') and accused['bail_status'] != 'none':
        cursor.close(); conn.close()
        return err('This accused already has an active bail. Revoke it first.', 400)

    fir_ids_raw = request.form.getlist('fir_ids') or request.form.get('fir_ids', '').split(',')
    fir_ids = [int(x) for x in fir_ids_raw if str(x).strip().isdigit()]
    valid_fir_ids = {f['id'] for f in arrest_firs}
    fir_ids = [fid for fid in dict.fromkeys(fir_ids) if fid in valid_fir_ids]
    if not fir_ids:
        cursor.close(); conn.close()
        return err('Select at least one valid arrest FIR.', 400)
    fir_id = fir_ids[0]

    bail_type = request.form.get('bail_type', 'temporary')
    bail_start = request.form.get('bail_start_date', '').strip() or None
    bail_end = request.form.get('bail_end_date', '').strip() or None
    bail_remark = request.form.get('bail_remark', '').strip()
    bail_rating = int(request.form.get('bail_rating', 0) or 0)
    if bail_type == 'permanent':
        bail_end = None

    today = datetime.now().date()
    try:
        start_date_obj = datetime.strptime(bail_start, '%Y-%m-%d').date() if bail_start else None
    except ValueError:
        start_date_obj = None
    if not start_date_obj or start_date_obj < today:
        cursor.close(); conn.close()
        return err('Bail start date must be today or in the future.', 400)
    if bail_end:
        try:
            end_date_obj = datetime.strptime(bail_end, '%Y-%m-%d').date()
        except ValueError:
            end_date_obj = None
        if not end_date_obj or end_date_obj < start_date_obj:
            cursor.close(); conn.close()
            return err('Bail end date cannot be before start date.', 400)

    photo_file = request.files.get('bail_photo')
    photo_lat_raw = request.form.get('bail_photo_lat', '').strip()
    photo_lng_raw = request.form.get('bail_photo_lng', '').strip()
    photo_time_raw = request.form.get('bail_photo_captured_at', '').strip()

    if not photo_file or not photo_file.filename:
        cursor.close(); conn.close()
        return err('A geo-tagged photo of the accused is required to approve bail.', 400)

    try:
        file_bytes = photo_file.read()
        mime_type = photo_file.mimetype or 'image/jpeg'
        b64_data = f"data:{mime_type};base64,{base64.b64encode(file_bytes).decode('utf-8')}"
        bail_photo_url, bail_photo_public_id = upload_image(b64_data, folder='accused_bail_photos')
    except Exception as e:
        logger.error(f"Bail photo upload error: {e}")
        bail_photo_url = None

    if not bail_photo_url:
        cursor.close(); conn.close()
        return err('Photo upload failed. Please try again.', 500)

    try:
        photo_lat_val = float(photo_lat_raw) if photo_lat_raw else None
    except ValueError:
        photo_lat_val = None
    try:
        photo_lng_val = float(photo_lng_raw) if photo_lng_raw else None
    except ValueError:
        photo_lng_val = None
    try:
        photo_time_val = datetime.fromisoformat(photo_time_raw) if photo_time_raw else datetime.now()
    except ValueError:
        photo_time_val = datetime.now()

    doc = request.files.get('bail_document')
    doc_url, doc_pub_id, doc_res_type = None, None, 'raw'
    if doc and doc.filename:
        doc_url, doc_pub_id, doc_res_type = upload_document(doc, folder='accused_bail_docs')

    set_as_profile_photo = not accused.get('photo_url')
    if set_as_profile_photo:
        cursor.execute("UPDATE accused_photos SET is_current=0 WHERE accused_id=%s", (accused_id,))
        cursor.execute("""INSERT INTO accused_photos (accused_id, photo_url, photo_public_id, is_current, uploaded_by)
                           VALUES (%s,%s,%s,1,%s)""", (accused_id, bail_photo_url, bail_photo_public_id, uid))

    cursor.execute("""
        UPDATE accused SET bail_status=%s, bail_start_date=%s, bail_end_date=%s,
        bail_documents_url=%s, bail_documents_public_id=%s,
        bail_photo_url=%s, bail_photo_public_id=%s,
        bail_photo_lat=%s, bail_photo_lng=%s, bail_photo_captured_at=%s,
        bail_remark=%s, bail_rating=%s, updated_by=%s
        {photo_cols}
        WHERE id=%s
    """.format(photo_cols=", photo_url=%s, photo_public_id=%s, profile_status='complete'" if set_as_profile_photo else ""),
    (bail_type, bail_start, bail_end, doc_url, doc_pub_id,
     bail_photo_url, bail_photo_public_id, photo_lat_val, photo_lng_val, photo_time_val,
     bail_remark, bail_rating, uid,
     *([bail_photo_url, bail_photo_public_id] if set_as_profile_photo else []),
     accused_id))

    cursor.execute("""
        INSERT INTO accused_bail_history
        (accused_id, fir_id, bail_type, bail_start_date, bail_end_date,
         bail_document_url, bail_document_public_id, bail_document_resource_type,
         bail_photo_url, bail_photo_public_id, bail_photo_lat, bail_photo_lng, bail_photo_captured_at,
         bail_remark, bail_rating, approved_by, status)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'ACTIVE')
    """, (accused_id, fir_id, bail_type, bail_start, bail_end,
          doc_url, doc_pub_id, doc_res_type or 'raw',
          bail_photo_url, bail_photo_public_id, photo_lat_val, photo_lng_val, photo_time_val,
          bail_remark, bail_rating, uid))
    bail_id = cursor.lastrowid

    for fid in fir_ids:
        cursor.execute("INSERT IGNORE INTO accused_bail_fir (bail_id, fir_id) VALUES (%s, %s)", (bail_id, fid))
    conn.commit()

    fir_rows_selected = [f for f in arrest_firs if f['id'] in fir_ids]
    fir_label = ', '.join(f"{f['fir_number']}/{f['thana']}" for f in fir_rows_selected) or '—'
    fir_row = fir_rows_selected[0] if fir_rows_selected else None

    log_activity(uid, role, f"Approved {bail_type} bail for accused ID:{accused_id} (FIR {fir_label})",
                 ip=request.remote_addr)

    try:
        from run import mail
    except Exception:
        mail = None
    try:
        send_bail_notification(
            district=fir_row['district'] if fir_row else session.get('district'),
            accused_name=accused['name'],
            fir_label=f"S/o {accused['fathers_name']} | FIR {fir_label}",
            bail_type=bail_type, bail_start=bail_start, bail_end=bail_end,
            bail_remark=bail_remark, bail_rating=bail_rating,
            approved_by_name=session.get('name'), approved_by_id=uid, mail_instance=mail,
            thana=fir_row['thana'] if fir_row else None)
    except Exception as e:
        logger.error(f"send_bail_notification error: {e}")

    cursor.close(); conn.close()
    return ok({'bail_id': bail_id, 'set_as_profile_photo': set_as_profile_photo})


@mobile_bp.route('/accused/<int:accused_id>/revoke-bail', methods=['POST'])
@role_required('admin')  # जमानत रद्द करना भी केवल जिला Admin का कार्य
def accused_revoke_bail(accused_id):
    data = request.get_json(silent=True) or request.form
    revoke_reason = (data.get('revoke_reason') or '').strip()
    uid = session['user_id']
    conn = get_connection(); cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        UPDATE accused_bail_history SET status='REVOKED', revoked_by=%s, revoked_at=NOW(), revoke_reason=%s
        WHERE accused_id=%s AND status='ACTIVE' ORDER BY approved_at DESC LIMIT 1
    """, (uid, revoke_reason or None, accused_id))
    cursor.execute("""
        UPDATE accused SET bail_status='none', bail_start_date=NULL, bail_end_date=NULL,
        bail_documents_url=NULL, bail_documents_public_id=NULL,
        bail_photo_url=NULL, bail_photo_public_id=NULL,
        bail_photo_lat=NULL, bail_photo_lng=NULL, bail_photo_captured_at=NULL,
        bail_remark=NULL, bail_rating=0, updated_by=%s WHERE id=%s
    """, (uid, accused_id))
    conn.commit(); cursor.close(); conn.close()
    log_activity(uid, session['role'], f"Revoked bail for accused ID:{accused_id}", ip=request.remote_addr)
    return ok()


@mobile_bp.route('/bailed-accused')
@role_required('admin', 'super_admin')
def bailed_accused():
    district = session.get('district')
    status_filter = request.args.get('status', 'ACTIVE')
    bail_type_filter = request.args.get('bail_type', '')
    search = request.args.get('search', '').strip()
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 25))

    conn = get_connection(); cursor = conn.cursor(dictionary=True)
    conditions = ["f.district=%s"]; params = [district]
    if status_filter in ('ACTIVE', 'REVOKED', 'COMPLETED'):
        conditions.append("abh.status=%s"); params.append(status_filter)
    if bail_type_filter in ('temporary', 'permanent'):
        conditions.append("abh.bail_type=%s"); params.append(bail_type_filter)
    if search:
        conditions.append("(a.name LIKE %s OR a.fathers_name LIKE %s)")
        like = f'%{search}%'; params += [like, like]
    where = " AND ".join(conditions)
    base_q = f"""
        SELECT abh.id AS bail_id, abh.bail_type, abh.bail_start_date, abh.bail_end_date,
               abh.bail_remark, abh.bail_rating, abh.status AS bail_history_status,
               abh.bail_document_url, abh.approved_at, abh.revoked_at, abh.completed_at,
               abh.revoke_reason, abh.photo_status, abh.source,
               a.id, a.name, a.fathers_name, a.photo_url,
               f.fir_number, f.thana, f.district,
               GROUP_CONCAT(DISTINCT CONCAT(f2.fir_number,'/',f2.thana)
                            ORDER BY f2.fir_number SEPARATOR ', ') AS all_firs,
               g.name AS approved_by_name, r.name AS revoked_by_name
        FROM accused_bail_history abh
        JOIN accused a ON a.id = abh.accused_id
        JOIN fir_cases f ON f.id = abh.fir_id
        LEFT JOIN accused_bail_fir abf ON abf.bail_id = abh.id
        LEFT JOIN fir_cases f2 ON f2.id = abf.fir_id
        LEFT JOIN users g ON g.id = abh.approved_by
        LEFT JOIN users r ON r.id = abh.revoked_by
        WHERE {where} GROUP BY abh.id ORDER BY abh.approved_at DESC
    """
    rows, total, total_pages = paginate_query(cursor, base_q, params, page, per_page)
    for b in rows:
        if not b.get('all_firs'):
            b['all_firs'] = f"{b['fir_number']}/{b['thana']}"
    cursor.close(); conn.close()
    return ok({'bail_records': rows_iso(rows), 'page': page, 'total': total,
               'total_pages': total_pages, 'per_page': per_page})


# ══════════════════════════════════════════════════════════════════════════
# FIR
# ══════════════════════════════════════════════════════════════════════════

@mobile_bp.route('/fir')
@role_required('admin', 'super_admin')
def fir_list():
    district = session.get('district')
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 25))
    search = request.args.get('search', '').strip()
    thana_f = request.args.get('thana', '').strip()
    fir_f = request.args.get('fir', '').strip()

    conditions = ["f.district=%s"]; params = [district]
    if thana_f:
        conditions.append("f.thana LIKE %s"); params.append(f'%{thana_f}%')
    if fir_f:
        conditions.append("f.fir_number LIKE %s"); params.append(f'%{fir_f}%')
    if search:
        conditions.append("(f.fir_number LIKE %s OR f.thana LIKE %s OR f.complainant LIKE %s OR f.acts LIKE %s)")
        like = f'%{search}%'; params += [like, like, like, like]
    where = " AND ".join(conditions)

    conn = get_connection(); cursor = conn.cursor(dictionary=True)
    base_q = f"""
        SELECT f.*, (SELECT COUNT(*) FROM accused_fir af WHERE af.fir_id=f.id) AS accused_count
        FROM fir_cases f WHERE {where} ORDER BY f.created_at DESC
    """
    rows, total, total_pages = paginate_query(cursor, base_q, params, page, per_page)
    cursor.execute("SELECT DISTINCT thana FROM fir_cases WHERE district=%s ORDER BY thana", (district,))
    thanas = [r['thana'] for r in cursor.fetchall()]
    cursor.close(); conn.close()
    return ok({'firs': rows_iso(rows), 'page': page, 'total': total,
               'total_pages': total_pages, 'per_page': per_page, 'thanas': thanas})


@mobile_bp.route('/fir/<int:fir_id>')
@role_required('admin', 'super_admin')
def fir_detail(fir_id):
    district = session.get('district')
    conn = get_connection(); cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM fir_cases WHERE id=%s AND district=%s", (fir_id, district))
    fir = cursor.fetchone()
    if not fir:
        cursor.close(); conn.close(); return err('FIR not found.', 404)
    cursor.execute("""
        SELECT a.id, a.name, a.fathers_name, a.photo_url, a.profile_status,
               af.in_total_accused, af.in_fir_accused, af.in_arrested, af.in_cs_accused,
               (SELECT COUNT(*) FROM accused_fir af2 WHERE af2.accused_id=a.id) AS fir_count
        FROM accused_fir af JOIN accused a ON a.id = af.accused_id
        WHERE af.fir_id = %s ORDER BY a.name
    """, (fir_id,))
    accused_list_rows = cursor.fetchall()
    cursor.close(); conn.close()
    return ok({'fir': jsonify_dates(dict(fir)), 'accused_list': rows_iso(accused_list_rows)})


@mobile_bp.route('/fir/add', methods=['POST'])
@role_required('admin', 'super_admin')
def fir_add():
    uid = session['user_id']; role = session['role']; district = session.get('district')
    data = request.get_json(silent=True) or request.form
    thana = (data.get('thana') or '').strip()
    fir_number = (data.get('fir_number') or '').strip()
    acts = (data.get('acts') or '').strip()
    total_raw = (data.get('total_accused') or '').strip()
    fir_raw = (data.get('fir_accused') or '').strip()
    arrested_raw = (data.get('arrested_accused') or '').strip()
    cs_raw = (data.get('cs_accused') or '').strip()
    complainant = (data.get('complainant') or '').strip()
    status = (data.get('status') or '').strip()

    if not thana or not fir_number:
        return err('Thana and FIR number are required.', 400)

    conn = get_connection(); cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("""
            INSERT INTO fir_cases (district, thana, fir_number, acts, total_accused_raw,
                fir_accused_raw, arrested_accused_raw, cs_accused_raw, complainant, status, created_by)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON DUPLICATE KEY UPDATE acts=VALUES(acts), complainant=VALUES(complainant),
                status=VALUES(status), total_accused_raw=VALUES(total_accused_raw),
                fir_accused_raw=VALUES(fir_accused_raw), arrested_accused_raw=VALUES(arrested_accused_raw),
                cs_accused_raw=VALUES(cs_accused_raw)
        """, (district, thana, fir_number, acts, total_raw, fir_raw, arrested_raw, cs_raw,
              complainant, status, uid))
        cursor.execute("SELECT id FROM fir_cases WHERE district=%s AND thana=%s AND fir_number=%s",
                       (district, thana, fir_number))
        fir_id = cursor.fetchone()['id']

        total_set = set(normalize_name(a['name']) for a in parse_accused_list(total_raw))
        fir_set = set(normalize_name(a['name']) for a in parse_accused_list(fir_raw))
        arrested_set = set(normalize_name(a['name']) for a in parse_accused_list(arrested_raw))
        cs_set = set(normalize_name(a['name']) for a in parse_accused_list(cs_raw))

        all_accused = parse_accused_list(total_raw)
        seen_norms = set(total_set)
        for a in parse_accused_list(fir_raw):
            n = normalize_name(a['name'])
            if n not in seen_norms:
                all_accused.append(a); seen_norms.add(n)

        linked_count = 0
        for acc in all_accused:
            acc_id, _created = find_or_create_accused(cursor, acc['name'], acc['fathers_name'], uid)
            if acc_id is None:
                continue
            n = normalize_name(acc['name'])
            upsert_accused_fir(cursor, acc_id, fir_id, n in total_set, n in fir_set,
                               n in arrested_set, n in cs_set)
            linked_count += 1

        cursor.execute("UPDATE fir_cases SET total_accused_count=%s WHERE id=%s", (linked_count, fir_id))
        conn.commit()
        log_activity(uid, role, f"FIR {fir_number}/{thana} manually created", ip=request.remote_addr)
        cursor.close(); conn.close()
        return ok({'fir_id': fir_id, 'linked_count': linked_count})
    except Exception as e:
        conn.rollback()
        logger.error(f"Manual FIR creation error: {e}")
        cursor.close(); conn.close()
        return err(str(e), 500)


# ══════════════════════════════════════════════════════════════════════════
# EXCEL BULK — FIR + accused upload, and sample download
# ══════════════════════════════════════════════════════════════════════════

@mobile_bp.route('/upload-accused', methods=['POST'])
@role_required('admin', 'super_admin')
def upload_accused():
    uid = session['user_id']; role = session['role']; district_session = session.get('district')
    file = request.files.get('excel_file')
    if not file or not file.filename:
        return err('Please choose an Excel file.', 400)
    fname = file.filename.lower()
    if not (fname.endswith('.xlsx') or fname.endswith('.xls')):
        return err('Only .xlsx or .xls files are supported.', 400)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
    except Exception as e:
        return err(f'Error reading Excel file: {e}', 400)

    conn = get_connection(); cursor = conn.cursor(dictionary=True)
    fir_success = accused_created = accused_linked = accused_skipped = 0
    failed = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not any(row):
            continue
        try:
            (sr_no, district, thana, fir_number, acts, total_raw, fir_raw, arrested_raw,
             cs_raw, complainant, status) = (list(row) + [None] * 11)[:11]
            district = str(district or '').strip() or district_session
            thana = str(thana or '').strip()
            fir_number = str(fir_number or '').strip()
            acts = str(acts or '').strip()
            total_raw = str(total_raw or '').strip()
            fir_raw = str(fir_raw or '').strip()
            arrested_raw = str(arrested_raw or '').strip()
            cs_raw = str(cs_raw or '').strip()
            complainant = str(complainant or '').strip()
            status = str(status or '').strip()

            if not fir_number or not thana:
                failed.append(f"Row {row_idx}: FIR number or thana empty"); continue

            cursor.execute("""
                INSERT INTO fir_cases (district, thana, fir_number, acts, total_accused_raw,
                    fir_accused_raw, arrested_accused_raw, cs_accused_raw, complainant, status, created_by)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON DUPLICATE KEY UPDATE acts=VALUES(acts), total_accused_raw=VALUES(total_accused_raw),
                    fir_accused_raw=VALUES(fir_accused_raw), arrested_accused_raw=VALUES(arrested_accused_raw),
                    cs_accused_raw=VALUES(cs_accused_raw), complainant=VALUES(complainant), status=VALUES(status)
            """, (district, thana, fir_number, acts, total_raw, fir_raw, arrested_raw, cs_raw,
                  complainant, status, uid))
            cursor.execute("SELECT id FROM fir_cases WHERE district=%s AND thana=%s AND fir_number=%s",
                           (district, thana, fir_number))
            fir_row = cursor.fetchone()
            if not fir_row:
                continue
            fir_id = fir_row['id']; fir_success += 1

            total_set = set(normalize_name(a['name']) for a in parse_accused_list(total_raw))
            fir_set = set(normalize_name(a['name']) for a in parse_accused_list(fir_raw))
            arrested_set = set(normalize_name(a['name']) for a in parse_accused_list(arrested_raw))
            cs_set = set(normalize_name(a['name']) for a in parse_accused_list(cs_raw))

            all_accused = parse_accused_list(total_raw)
            seen_norms = set(total_set)
            for a in parse_accused_list(fir_raw):
                n = normalize_name(a['name'])
                if n not in seen_norms:
                    all_accused.append(a); seen_norms.add(n)

            linked_this_fir = 0
            for acc in all_accused:
                acc_norm = normalize_name(acc['name'])
                accused_id, was_created = find_or_create_accused(cursor, acc['name'], acc['fathers_name'], uid)
                if accused_id is None:
                    accused_skipped += 1; continue
                if was_created:
                    accused_created += 1
                upsert_accused_fir(cursor, accused_id, fir_id, acc_norm in total_set, acc_norm in fir_set,
                                   acc_norm in arrested_set, acc_norm in cs_set)
                accused_linked += 1; linked_this_fir += 1

            cursor.execute("UPDATE fir_cases SET total_accused_count=%s WHERE id=%s", (linked_this_fir, fir_id))
        except Exception as e:
            logger.error(f"Row {row_idx} error: {e}")
            failed.append(f"Row {row_idx}: {str(e)[:100]}")
            continue

    conn.commit(); cursor.close(); conn.close()
    log_activity(uid, role,
                 f"Excel upload: {fir_success} FIR, {accused_created} new accused, "
                 f"{accused_linked} links, {accused_skipped} skipped", ip=request.remote_addr)
    return ok({'fir_success': fir_success, 'accused_created': accused_created,
               'accused_linked': accused_linked, 'accused_skipped': accused_skipped,
               'failed': failed})


@mobile_bp.route('/download-accused-sample')
@role_required('admin', 'super_admin')
def download_accused_sample():
    return download_accused_sample_file()


# ══════════════════════════════════════════════════════════════════════════
# BAIL EXCEL BULK-APPROVAL (court order lists)
# ══════════════════════════════════════════════════════════════════════════

@mobile_bp.route('/bail-excel/upload', methods=['POST'])
@role_required('admin', 'super_admin')
def bail_excel_upload():
    district = session.get('district')
    file = request.files.get('excel_file')
    if not file or not file.filename:
        return err('Please choose a file.', 400)
    try:
        batch_id = stage_bail_excel(file, file.filename, district, session['user_id'])
        return ok({'batch_id': batch_id})
    except Exception as e:
        logger.error(f"stage_bail_excel error: {e}")
        return err(str(e), 500)


@mobile_bp.route('/bail-excel/batches')
@role_required('admin', 'super_admin')
def bail_excel_batches():
    district = session.get('district')
    return ok({'batches': rows_iso(list_batches(district))})


@mobile_bp.route('/bail-excel/batch/<int:batch_id>')
@role_required('admin', 'super_admin')
def bail_excel_batch_review(batch_id):
    district = session.get('district')
    batch, rows = get_batch_review(batch_id, district)
    if not batch:
        return err('Batch not found.', 404)
    return ok({'batch': jsonify_dates(dict(batch)), 'rows': rows_iso(rows),
               'status_labels': MATCH_STATUS_LABELS})


@mobile_bp.route('/bail-excel/batch/<int:batch_id>/row/<int:row_id>/resolve', methods=['POST'])
@role_required('admin', 'super_admin')
def bail_excel_resolve(batch_id, row_id):
    district = session.get('district')
    data = request.get_json(silent=True) or request.form
    accused_id = int(data.get('accused_id', 0) or 0)
    success, message = resolve_ambiguous_row(batch_id, row_id, accused_id, district)
    if not success:
        return err(message, 400)
    return ok(message=message)


@mobile_bp.route('/bail-excel/batch/<int:batch_id>/confirm', methods=['POST'])
@role_required('admin', 'super_admin')
def bail_excel_confirm(batch_id):
    district = session.get('district')
    data = request.get_json(silent=True) or {}
    row_ids = [int(x) for x in data.get('row_ids', [])]
    result = confirm_batch(batch_id, district, row_ids, session['user_id'])
    if result.get('error'):
        return err(result['error'], 400)
    log_activity(session['user_id'], session['role'],
                 f"Confirmed bail-excel batch {batch_id}: {len(result['created'])} created",
                 ip=request.remote_addr)
    return ok(result)


@mobile_bp.route('/bail-excel/batch/<int:batch_id>/discard', methods=['POST'])
@role_required('admin', 'super_admin')
def bail_excel_discard(batch_id):
    district = session.get('district')
    success, message = discard_batch(batch_id, district, session['user_id'])
    if not success:
        return err(message, 400)
    return ok(message=message)


@mobile_bp.route('/bail-pending-photos')
@role_required('admin', 'super_admin')
def bail_pending_photos():
    district = session.get('district')
    return ok({'pending': rows_iso(list_pending_photo_bails(district))})


@mobile_bp.route('/bail-pending-photos/<int:bail_id>/complete', methods=['POST'])
@role_required('admin')  # फ़ोटो/दस्तावेज़ पूर्ण करना केवल जिला Admin का कार्य; super_admin केवल list देख सकता है
def bail_complete_photo(bail_id):
    import base64
    district = session.get('district')
    photo_file = request.files.get('photo')
    doc_file = request.files.get('document')
    photo_data = None
    if photo_file and photo_file.filename:
        file_bytes = photo_file.read()
        mime_type = photo_file.mimetype or 'image/jpeg'
        photo_data = f"data:{mime_type};base64,{base64.b64encode(file_bytes).decode('utf-8')}"
    success, message = complete_bail_photo(bail_id, district, session['user_id'],
                                            photo_data=photo_data, doc_file=doc_file)
    if not success:
        return err(message, 400)
    return ok(message=message)


# ══════════════════════════════════════════════════════════════════════════
# MASTER — super admin management, all admins, activity logs
# ══════════════════════════════════════════════════════════════════════════

@mobile_bp.route('/master/super-admins')
@role_required('master')
def master_super_admins():
    conn = get_connection(); cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT u.*, (SELECT COUNT(*) FROM users a WHERE a.created_by=u.id AND a.role='admin') as admin_count
        FROM users u WHERE u.role='super_admin' ORDER BY u.created_at DESC
    """)
    rows = cursor.fetchall()
    cursor.close(); conn.close()
    return ok({'supers': rows_iso(rows)})


@mobile_bp.route('/master/create-super-admin', methods=['POST'])
@role_required('master')
def master_create_super_admin():
    data = request.get_json(silent=True) or request.form
    fields = {k: (data.get(k) or '').strip() for k in
              ['user_id', 'name', 'designation', 'contact', 'email', 'district', 'address', 'password']}
    if not all([fields['user_id'], fields['name'], fields['email'], fields['password']]):
        return err('All required fields must be filled.', 400)

    conn = get_connection(); cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT id FROM users WHERE user_id=%s OR email=%s", (fields['user_id'], fields['email']))
    if cursor.fetchone():
        cursor.close(); conn.close(); return err('User ID or Email already exists.', 409)
    cursor.execute("""
        INSERT INTO users (user_id,name,designation,contact,email,district,address,password_hash,role,created_by,is_active)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,'super_admin',%s,1)
    """, (fields['user_id'], fields['name'], fields['designation'], fields['contact'],
          fields['email'], fields['district'], fields['address'],
          generate_password_hash(fields['password']), session['user_id']))
    conn.commit()
    log_activity(session['user_id'], 'master', f"Created super admin: {fields['user_id']}", ip=request.remote_addr)
    cursor.close(); conn.close()
    return ok(message=f"Super Admin '{fields['name']}' created successfully.")


@mobile_bp.route('/master/edit-user/<int:uid>', methods=['GET', 'POST'])
@role_required('master')
def master_edit_user(uid):
    conn = get_connection(); cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM users WHERE id=%s AND role!='master'", (uid,))
    user = cursor.fetchone()
    if not user:
        cursor.close(); conn.close(); return err('User not found.', 404)

    if request.method == 'GET':
        cursor.close(); conn.close()
        return ok({'user': jsonify_dates(dict(user))})

    data = request.get_json(silent=True) or request.form
    name = (data.get('name') or '').strip()
    designation = (data.get('designation') or '').strip()
    contact = (data.get('contact') or '').strip()
    email = (data.get('email') or '').strip()
    district = (data.get('district') or '').strip()
    address = (data.get('address') or '').strip()
    cursor.execute("""UPDATE users SET name=%s, designation=%s, contact=%s, email=%s,
                       district=%s, address=%s WHERE id=%s""",
                   (name, designation, contact, email, district, address, uid))
    conn.commit()
    log_activity(session['user_id'], 'master', f"Edited user ID: {uid}", ip=request.remote_addr)
    cursor.close(); conn.close()
    return ok(message='User updated successfully.')


@mobile_bp.route('/master/revoke-user/<int:uid>', methods=['POST'])
@role_required('master')
def master_revoke_user(uid):
    conn = get_connection(); cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM users WHERE id=%s AND role!='master'", (uid,))
    user = cursor.fetchone()
    if not user:
        cursor.close(); conn.close(); return err('User not found.', 404)
    new_status = 0 if user['is_active'] else 1
    cursor.execute("UPDATE users SET is_active=%s WHERE id=%s", (new_status, uid))
    conn.commit()
    action = 'Revoked' if new_status == 0 else 'Restored'
    log_activity(session['user_id'], 'master', f"{action} user: {user['user_id']}", ip=request.remote_addr)
    cursor.close(); conn.close()
    return ok({'is_active': bool(new_status)}, message=f"User access {action.lower()} successfully.")


@mobile_bp.route('/master/all-admins')
@role_required('master')
def master_all_admins():
    search = request.args.get('search', ''); page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 20))
    conn = get_connection(); cursor = conn.cursor(dictionary=True)
    base_q = """
        SELECT a.*, s.name as super_name, s.user_id as super_uid, s.district as super_district
        FROM users a LEFT JOIN users s ON a.created_by = s.id WHERE a.role='admin'
    """
    params = []
    if search:
        base_q += " AND (a.name LIKE %s OR a.user_id LIKE %s OR a.district LIKE %s)"
        params += [f'%{search}%', f'%{search}%', f'%{search}%']
    base_q += " ORDER BY a.created_at DESC"
    rows, total, total_pages = paginate_query(cursor, base_q, params, page, per_page)
    cursor.close(); conn.close()
    return ok({'admins': rows_iso(rows), 'page': page, 'total': total, 'total_pages': total_pages})


@mobile_bp.route('/master/logs')
@role_required('master')
def master_logs():
    page = int(request.args.get('page', 1)); per_page = int(request.args.get('per_page', 50))
    role_filter = request.args.get('role', ''); search = request.args.get('search', '')
    conn = get_connection(); cursor = conn.cursor(dictionary=True)
    base_q = """
        SELECT l.*, u.name as user_name, u.user_id as uid
        FROM activity_logs l LEFT JOIN users u ON l.user_id = u.id WHERE 1=1
    """
    params = []
    if role_filter:
        base_q += " AND l.user_role=%s"; params.append(role_filter)
    if search:
        base_q += " AND (l.action LIKE %s OR l.endpoint LIKE %s OR u.name LIKE %s)"
        params += [f'%{search}%', f'%{search}%', f'%{search}%']
    base_q += " ORDER BY l.created_at DESC"
    rows, total, total_pages = paginate_query(cursor, base_q, params, page, per_page)
    cursor.close(); conn.close()
    return ok({'logs': rows_iso(rows), 'page': page, 'total': total, 'total_pages': total_pages})


# ══════════════════════════════════════════════════════════════════════════
# SUPER ADMIN — district admin management
# ══════════════════════════════════════════════════════════════════════════

@mobile_bp.route('/super/admins')
@role_required('super_admin')
def super_admins():
    search = request.args.get('search', '').strip()
    page = int(request.args.get('page', 1)); per_page = int(request.args.get('per_page', 20))
    conn = get_connection(); cursor = conn.cursor(dictionary=True)
    base_q = """
        SELECT u.*, c.name AS created_by_name FROM users u
        LEFT JOIN users c ON c.id = u.created_by
        WHERE u.created_by=%s AND u.role='admin'
    """
    params = [session['user_id']]
    if search:
        base_q += " AND (u.name LIKE %s OR u.user_id LIKE %s OR u.designation LIKE %s OR u.contact LIKE %s)"
        like = f'%{search}%'; params += [like, like, like, like]
    base_q += " ORDER BY u.created_at DESC"
    rows, total, total_pages = paginate_query(cursor, base_q, params, page, per_page)
    cursor.close(); conn.close()
    return ok({'admins': rows_iso(rows), 'page': page, 'total': total, 'total_pages': total_pages})


@mobile_bp.route('/super/create-admin', methods=['POST'])
@role_required('super_admin')
def super_create_admin():
    data = request.get_json(silent=True) or request.form
    uid = (data.get('user_id') or '').strip().upper()
    name = (data.get('name') or '').strip()
    designation = (data.get('designation') or '').strip()
    contact = (data.get('contact') or '').strip()
    email = (data.get('email') or '').strip()
    district = session.get('district')
    address = (data.get('address') or '').strip()
    password = (data.get('password') or '').strip()

    if not all([uid, name, email, password]):
        return err('User ID, name, email and password are required.', 400)

    conn = get_connection(); cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT id FROM users WHERE user_id=%s OR email=%s", (uid, email))
    if cursor.fetchone():
        cursor.close(); conn.close(); return err('User ID or email already exists.', 409)
    cursor.execute("""
        INSERT INTO users (user_id,name,designation,contact,email,district,address,password_hash,role,created_by,is_active)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,'admin',%s,1)
    """, (uid, name, designation, contact, email, district, address,
          generate_password_hash(password), session['user_id']))
    conn.commit()
    log_activity(session['user_id'], session['role'], f"Created admin {uid}", ip=request.remote_addr)
    cursor.close(); conn.close()
    return ok(message=f'Admin {name} ({uid}) created successfully.')


@mobile_bp.route('/super/admin/<int:uid>/toggle', methods=['POST'])
@role_required('super_admin')
def super_toggle_admin(uid):
    conn = get_connection(); cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT id,is_active,user_id FROM users WHERE id=%s AND created_by=%s",
                   (uid, session['user_id']))
    user = cursor.fetchone()
    if not user:
        cursor.close(); conn.close(); return err('Admin not found.', 404)
    new_status = 0 if user['is_active'] else 1
    cursor.execute("UPDATE users SET is_active=%s WHERE id=%s", (new_status, uid))
    conn.commit(); cursor.close(); conn.close()
    return ok({'is_active': bool(new_status)})


@mobile_bp.route('/super/upload-admins', methods=['POST'])
@role_required('super_admin')
def super_upload_admins():
    import csv, io as _io
    file = request.files.get('csv_file')
    if not file or not file.filename.endswith('.csv'):
        return err('Please upload a CSV file.', 400)
    stream = _io.StringIO(file.stream.read().decode('utf-8'))
    reader = csv.DictReader(stream)
    conn = get_connection(); cursor = conn.cursor(dictionary=True)
    success = 0; failed = []
    for i, row in enumerate(reader, start=2):
        uid = (row.get('user_id', '') or '').strip().upper()
        name = (row.get('name', '') or '').strip()
        email = (row.get('email', '') or '').strip()
        pwd = (row.get('password', '') or '').strip()
        if not all([uid, name, email, pwd]):
            failed.append(f"Row {i}: required field empty"); continue
        cursor.execute("SELECT id FROM users WHERE user_id=%s OR email=%s", (uid, email))
        if cursor.fetchone():
            failed.append(f"Row {i} ({uid}): already exists"); continue
        try:
            cursor.execute("""
                INSERT INTO users (user_id,name,designation,contact,email,district,address,password_hash,role,created_by,is_active)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,'admin',%s,1)
            """, (uid, name, (row.get('designation', '') or '').strip(), (row.get('contact', '') or '').strip(),
                  email, session.get('district'), (row.get('address', '') or '').strip(),
                  generate_password_hash(pwd), session['user_id']))
            success += 1
        except Exception as e:
            failed.append(f"Row {i}: {str(e)[:80]}")
    conn.commit(); cursor.close(); conn.close()
    return ok({'success_count': success, 'failed': failed})


@mobile_bp.route('/super/download-admin-sample')
@role_required('super_admin')
def super_download_admin_sample():
    import csv, io as _io
    output = _io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["user_id", "name", "designation", "contact", "email", "address", "password"])
    writer.writerow(["ADM001", "Rahul Sharma", "Head Constable", "9876543210",
                      "rahul@example.com", "Police Station Mirzapur", "Admin@123"])
    buf = io.BytesIO(output.getvalue().encode('utf-8'))
    return send_file(buf, mimetype='text/csv', as_attachment=True, download_name='admin_sample.csv')


# ══════════════════════════════════════════════════════════════════════════
# थाना (POLICE STATION) DIRECTORY — mirrors web super_admin.py /thana routes.
# WhatsApp/email bail-approval notifications are routed through this list.
# ══════════════════════════════════════════════════════════════════════════

@mobile_bp.route('/super/thana')
@role_required('super_admin')
def super_thana_list():
    from thana_service import list_thanas
    search = request.args.get('search', '').strip()
    thanas = list_thanas(session.get('district'), search)
    return ok({'thanas': rows_iso(thanas)})


@mobile_bp.route('/super/thana/add', methods=['POST'])
@role_required('super_admin')
def super_thana_add():
    from thana_service import add_thana
    data = request.get_json(silent=True) or request.form
    okd, error = add_thana(
        session.get('district'),
        (data.get('thana_name') or '').strip(),
        (data.get('contact') or '').strip(),
        (data.get('email') or '').strip(),
        session['user_id'],
    )
    if not okd:
        return err(error or 'थाना जोड़ने में त्रुटि।', 400)
    log_activity(session['user_id'], session['role'], f"Added/updated थाना {data.get('thana_name')}",
                 ip=request.remote_addr)
    return ok()


@mobile_bp.route('/super/thana/<int:thana_id>/edit', methods=['POST'])
@role_required('super_admin')
def super_thana_edit(thana_id):
    from thana_service import update_thana
    data = request.get_json(silent=True) or request.form
    okd, error = update_thana(
        thana_id, session.get('district'),
        (data.get('thana_name') or '').strip(),
        (data.get('contact') or '').strip(),
        (data.get('email') or '').strip(),
    )
    if not okd:
        return err(error or 'अपडेट करने में त्रुटि।', 400)
    return ok()


@mobile_bp.route('/super/thana/<int:thana_id>/toggle', methods=['POST'])
@role_required('super_admin')
def super_thana_toggle(thana_id):
    from thana_service import toggle_thana
    if not toggle_thana(thana_id, session.get('district')):
        return err('थाना नहीं मिला।', 404)
    return ok()


@mobile_bp.route('/super/thana/<int:thana_id>/delete', methods=['POST'])
@role_required('super_admin')
def super_thana_delete(thana_id):
    from thana_service import delete_thana
    if not delete_thana(thana_id, session.get('district')):
        return err('थाना नहीं मिला।', 404)
    log_activity(session['user_id'], session['role'], f"Deleted थाना #{thana_id}", ip=request.remote_addr)
    return ok()


@mobile_bp.route('/super/thana/upload', methods=['POST'])
@role_required('super_admin')
def super_thana_upload():
    from thana_service import bulk_upload_thanas_from_excel
    file = request.files.get('excel_file')
    if not file or not file.filename:
        return err('कृपया Excel फ़ाइल चुनें।', 400)
    fname = file.filename.lower()
    if not (fname.endswith('.xlsx') or fname.endswith('.xls')):
        return err('केवल .xlsx या .xls फ़ाइल अपलोड करें।', 400)
    result = bulk_upload_thanas_from_excel(session.get('district'), file, session['user_id'])
    log_activity(session['user_id'], session['role'],
                 f"Bulk uploaded {result['success']} थाना", ip=request.remote_addr)
    return ok({'success_count': result['success'], 'failed': result['failed']})


@mobile_bp.route('/super/thana/download-sample')
@role_required('super_admin')
def super_thana_download_sample():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'थाना सूची'
    ws.append(['thana_name', 'contact', 'email'])
    header_fill = PatternFill("solid", fgColor="1a73e8")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    ws.append(['थाना कोतवाली', '9876543210', 'kotwali@example.com'])
    ws.append(['थाना देहात', '9876500000', 'dehat.thana@example.com'])
    for i, w in enumerate([25, 18, 30], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                      as_attachment=True, download_name='thana_upload_sample.xlsx')
